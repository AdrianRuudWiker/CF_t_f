# -*- coding: utf-8 -*-
"""
Regenererer MC-motoren i Kontantstromsmodell_petroleum.xlsx fra input, med
forventningsforankring av simuleringen i NB26-basis (K3-fiksen).

Bakgrunn: prisen var medianforankret i en høyreskjev lognormal, slik at
E[pris] > NB26 og simuleringens forventning drev systematisk over NB26s eget
anslag. NB26-banen er en forventningsbane, ikke en median. Fiksen forankrer
forventningen:

  Pris  (Jensen): gang prisfaktoren med EXP(-0,5*Var_t), der
                  Var_t = sigma^2 * (1 - phi^(2t)) / (1 - phi^2), phi = 1-kappa.
                  Korreksjonen ligger KUN på eksponentieringen, ikke i OU-
                  rekursjonen for logMo/logMg (ellers reverterer prosessen mot
                  et senket nivå og feilen komponerer).
  Volum (mean-1): del volumfaktoren på sin forventning 1 + (fh + fl - 2)/6
                  (lukket form for den symmetriske triangulære vekten).

Etter fiksen er E[simulering] = NB26-basis i både pris og volum.

Denne runden (bare K3-fiksen): MC-motoren regenereres fullstendig fra input;
de synlige arkene og formateringen bæres uendret fra malen (dagens bok), og
to utdaterte forklaringsceller i Dokumentasjon/Monte Carlo synkroniseres.
De faste trekkene (w, z1, z2, frø 2026) er en fast del av inputen og bevares
verbatim, slik at eneste endring mot dagens bok er selve forankringen.

Kjøring:  python3 build_workbook.py
Krever:   numpy, openpyxl
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as CL

TEMPLATE = "Kontantstromsmodell_petroleum.xlsx"
OUT = "Kontantstromsmodell_petroleum.xlsx"

YEARS = list(range(2026, 2051))          # 25 modellår
NY = len(YEARS)
FU_ROW0 = 18                             # Forutsetninger: 2026 ligger på rad 18
DATA_ROW0 = 10                           # MC-motor: første simulering på rad 10

# Kolonnebaser i MC-motoren (1-indeksert)
C_SIM, C_W = 1, 2
C_Z1, C_Z2 = 3, 28
C_LMO, C_LMG, C_SNCF = 53, 78, 103
C_NPV3, C_NPV2, C_NPV4, C_KUM = 128, 129, 130, 131


def build(template=TEMPLATE, out=OUT, n=2000, draws=None, doc_sync=True):
    wb = load_workbook(template)
    ws = wb["MC-motor"]

    # ---- Input: faste trekk (bevar fra malen, eller bruk oppgitte) ----
    if draws is None:
        w = [ws.cell(DATA_ROW0 + r, C_W).value for r in range(n)]
        z1 = [[ws.cell(DATA_ROW0 + r, C_Z1 + i).value for i in range(NY)] for r in range(n)]
        z2 = [[ws.cell(DATA_ROW0 + r, C_Z2 + i).value for i in range(NY)] for r in range(n)]
    else:
        w, z1, z2 = draws
    assert len(w) == n

    # ---- Topptekst og etiketter ----
    ws.cell(1, 1, f"MC-MOTOR: {n} simuleringer (forventningsforankret)")
    for r, lbl in zip(range(3, 9), ["A = volO*pO + volN*pN", "B = volG*pG",
                                    "C = kostnader", "andel", "fh", "fl"]):
        ws.cell(r, 2, lbl)
    ws.cell(1, 2, "VarO_t")
    ws.cell(2, 2, "VarG_t")
    hdr = ["sim", "w"] + [f"z1_{y}" for y in YEARS] + [f"z2_{y}" for y in YEARS] \
        + [f"lMo_{y}" for y in YEARS] + [f"lMg_{y}" for y in YEARS] \
        + [f"SNCF_{y}" for y in YEARS] + ["NPV3", "NPV2", "NPV4", "KUM"]
    for c, h in enumerate(hdr, 1):
        ws.cell(9, c, h)

    # ---- Per-år parameterblokk (kol CY..DW) + varians (rad 1-2) ----
    for i in range(NY):
        col = C_SNCF + i
        cl = CL(col)
        fr = FU_ROW0 + i                 # Forutsetninger-rad for året
        t = i + 1                        # årsindeks 1..25 for variansen
        # Løpende varians (lukket form), phi = 1-kappa
        ws.cell(1, col, f"=Forutsetninger!$B$11^2*(1-(1-Forutsetninger!$B$14)^(2*{t}))"
                        f"/(1-(1-Forutsetninger!$B$14)^2)")
        ws.cell(2, col, f"=Forutsetninger!$B$12^2*(1-(1-Forutsetninger!$B$15)^(2*{t}))"
                        f"/(1-(1-Forutsetninger!$B$15)^2)")
        ws.cell(3, col, f"=Forutsetninger!B{fr}*Forutsetninger!J{fr}"
                        f"+Forutsetninger!D{fr}*Forutsetninger!L{fr}")
        ws.cell(4, col, f"=Forutsetninger!C{fr}*Forutsetninger!K{fr}")
        ws.cell(5, col, f"=Forutsetninger!M{fr}")
        ws.cell(6, col, f"=Forutsetninger!P{fr}")
        ws.cell(7, col, f"=Forutsetninger!H{fr}")
        ws.cell(8, col, f"=Forutsetninger!I{fr}")

    # ---- Simuleringsrader ----
    for r in range(n):
        row = DATA_ROW0 + r
        ws.cell(row, C_SIM, r + 1)
        ws.cell(row, C_W, w[r])
        for i in range(NY):
            ws.cell(row, C_Z1 + i, z1[r][i])
            ws.cell(row, C_Z2 + i, z2[r][i])
        # logMo (OU på logpris) — uendret rekursjon
        for i in range(NY):
            col = C_LMO + i
            z1c = CL(C_Z1 + i)
            if i == 0:
                f = f"=Forutsetninger!$B$11*{z1c}{row}"
            else:
                prev = CL(C_LMO + i - 1)
                f = f"=(1-Forutsetninger!$B$14)*{prev}{row}+Forutsetninger!$B$11*{z1c}{row}"
            ws.cell(row, col, f)
        # logMg (OU, korrelasjon i formelen)
        for i in range(NY):
            col = C_LMG + i
            z1c, z2c = CL(C_Z1 + i), CL(C_Z2 + i)
            shock = (f"Forutsetninger!$B$12*(Forutsetninger!$B$13*{z1c}{row}"
                     f"+SQRT(1-Forutsetninger!$B$13^2)*{z2c}{row})")
            if i == 0:
                f = f"={shock}"
            else:
                prev = CL(C_LMG + i - 1)
                f = f"=(1-Forutsetninger!$B$15)*{prev}{row}+{shock}"
            ws.cell(row, col, f)
        # SNCF — FORVENTNINGSFORANKRET (K3-fiks)
        for i in range(NY):
            col = C_SNCF + i
            cl = CL(col)
            lmo, lmg = CL(C_LMO + i), CL(C_LMG + i)
            volfac = f"IF($B{row}>=0,1+$B{row}*({cl}$7-1),1+$B{row}*(1-{cl}$8))"
            volmean = f"(1+({cl}$7+{cl}$8-2)/6)"          # E[volfac_t]
            rev = (f"({cl}$3*EXP({lmo}{row}-0.5*{cl}$1)"
                   f"+{cl}$4*EXP({lmg}{row}-0.5*{cl}$2)-{cl}$5)")
            ws.cell(row, col, f"={cl}$6*{volfac}/{volmean}*{rev}/1000")
        # NPV/KUM — uendret
        rng = f"{CL(C_SNCF)}{row}:{CL(C_SNCF+NY-1)}{row}"
        ws.cell(row, C_NPV3, f"=NPV(Forutsetninger!$B$8,{rng})")
        ws.cell(row, C_NPV2, f"=NPV(Forutsetninger!$B$9,{rng})")
        ws.cell(row, C_NPV4, f"=NPV(Forutsetninger!$B$10,{rng})")
        ws.cell(row, C_KUM, f"=SUM({rng})")

    if doc_sync:
        _doc_sync(wb)

    # openpyxl fjerner bufrede formelverdier ved lagring; tving full
    # reberegning når boka åpnes, så ingen celle står tom i manuell kalk-modus.
    wb.calculation.fullCalcOnLoad = True

    wb.save(out)
    return out


def _doc_sync(wb):
    """Minimal synk av forklaringsceller som ble utdatert av forankringsfiksen."""
    dok = wb["Dokumentasjon"]
    dok["A13"] = (
        "5. Forventningsforankring: NB26-prisbanen er FORVENTNINGSBANEN, og "
        "simuleringen er forankret slik at forventet SNCF er lik NB26-basis i "
        "både pris og volum (E[pris]=NB26 via Jensen-korreksjon på logpris; "
        "E[volumfaktor]=1). Pga. lognormal høyreskjevhet ligger medianbanen "
        "noe under forventningen - det er statistikk, ikke feil.")
    # Estimeringsusikkerhet i sigma/kappa (antakelse #7) - punktestimater brukt
    # som sentrale, usikkerheten dokumentert (jf. brukervalg).
    a15 = dok["A15"].value
    if "relativ standardfeil" not in a15:
        dok["A15"] = a15 + (
            " Standardavvikene (sigma_olje 0,2325 / sigma_gass 0,3856) og "
            "reverteringsfarten (kappa) er dermed estimert paa faa "
            "observasjoner: sigma har om lag 14 pst. relativ standardfeil, og "
            "kappa er mer usikker og nedadbiaset paa kort utvalg. "
            "Punktestimatene brukes som sentrale; usikkerheten tilsier en "
            "foelsomhet paa sigma/kappa som egen kjoering (utsatt).")
    mc = wb["Monte Carlo"]
    mc["A2"] = mc["A2"].value.replace("medianforankret", "forventningsforankret")
    mc["B47"] = "forventning = NB26-banen (E[pris]=NB26)"


if __name__ == "__main__":
    p = build()
    print("Skrev", p)
