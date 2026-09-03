"""Kartlegger strukturen i kildefilene FØR det skrives uttrekkskode.

Dette skriptet er obligatorisk første steg. Radnumrene i README er en hypotese
hentet fra gjennomgangen av den forrige modellen, ikke en spesifikasjon. Den
forrige modellen ble skrevet mot antatte radnumre, og resultatet var en lavbane
som lå over basisbanen i fire år.

Kjør:
    python -m src.inspiser
"""

from __future__ import annotations

import glob
import os

import openpyxl

RAW = "data/raw"
MAKS_RAD = 120


def _tallspenn(row):
    tall = [c for c in row if isinstance(c, (int, float)) and not isinstance(c, bool)]
    if not tall:
        return "", 0
    return f"{tall[0]:>13,.1f} ... {tall[-1]:>13,.1f}", len(tall)


def _etikett(row, n=5):
    for c in row[:n]:
        if isinstance(c, str) and c.strip():
            return c.strip()[:48]
    return ""


def vis_ark(ws, maks_rad: int = MAKS_RAD) -> None:
    print(f"\n--- ark: {ws.title}   ({ws.max_row} rader x {ws.max_column} kolonner)")
    print(f"{'rad':>4} | {'etikett':48s} | {'n':>4} | spenn")
    print("-" * 100)
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > maks_rad:
            print(f"     ... avkortet ved rad {maks_rad}")
            break
        etikett = _etikett(row)
        spenn, n = _tallspenn(row)
        if etikett or n:
            print(f"{i:4d} | {etikett:48s} | {n:4d} | {spenn}")


def vis_fil(sti: str, ark: list[str] | None = None) -> None:
    print("\n" + "=" * 100)
    print(f"FIL: {os.path.basename(sti)}")
    print("=" * 100)
    wb = openpyxl.load_workbook(sti, data_only=True, read_only=True)
    navn = [w.title for w in wb.worksheets]
    print(f"ark: {navn}")
    for a in (ark or navn):
        if a in navn:
            vis_ark(wb[a])
        else:
            print(f"\n!! fant ikke arket {a!r}")
    wb.close()


def main() -> None:
    filer = sorted(glob.glob(f"{RAW}/*.xlsx"))
    if not filer:
        print(f"Ingen .xlsx i {RAW}/. Legg kildefilene der først.")
        return

    for sti in filer:
        navn = os.path.basename(sti).lower()
        if "mulighetsbilde" in navn:
            vis_fil(sti, ["Formue", "Skiftberegning", "KVARTS"])
        else:
            vis_fil(sti)

    print("\n" + "=" * 100)
    print("SE ETTER:")
    print("  - hvilken rad hver etikett faktisk ligger på")
    print("  - hvilken rad årstallene står i, og hvilken kolonne serien starter i")
    print("  - størrelsesorden på første og siste verdi -> mill. eller mrd.?")
    print("  - hvilket år hvert ark starter (KVARTS 1997, Formue 2007, Sodir 1970)")
    print("\nIKKE skriv uttrekkskode før dette er avklart med bruker.")


if __name__ == "__main__":
    main()
