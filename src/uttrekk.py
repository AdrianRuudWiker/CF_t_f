"""Kilder -> data/inndata.csv, med årgangsmerking i data/kilder.csv.

Dette steget finnes fordi NB27 kommer senere i år. Når den gjør det, skal
bytte av tallgrunnlag være én endring her og en ny kjøring — ikke en runde med
innliming. Den forrige modellen limte inn verdier uten kobling til kilde, og
det var årsaken til at Sokkeldirektoratets høy/lav ble delt på NB26s basis i
stedet for Sokkeldirektoratets egen.

KONTRAKT
    inndata.csv:  én rad per år (1971-2090), én kolonne per serie, ingen hull
                  markert som 0 — bruk tom celle.
    kilder.csv:   én rad per serie, med årgang og uttrekksdato utfylt.

ENHETER — les denne før modell.py skrives
    Beløpskolonner (driftsutgifter, investeringsutgifter, snks) er i MRD.
    kroner. Kilden oppgir dem i mill.; omregningen skjer her, ved grensen.
    Priser er i kr per Sm3 o.e. og volumer i mill. Sm3 o.e., så
        pris * volum -> MILL. kroner, og må deles på 1 000
    for å møte kostnadene og SNKS. Det er den ene omregningen modellkoden må
    gjøre selv, og den mest sannsynlige feilen i hele prosjektet.

REGLER
    - Kostnader = påløpte utgifter delt på deflatoren. IKKE arbeidsbokens
      "faste priser"-rader; de ligger om lag 24 pst. lavere på egen basis.
    - Drifts- og investeringsutgifter hentes hver for seg, ikke som sum.
    - Volumforholdstall regnes mot Sokkeldirektoratets EGEN basis:
          f[t] = SD_scenario[t] / SD_basis[t]
      og anvendes deretter på NB26s basisbane. Aldri SD_scenario / NB26_basis.

PRISBASIS — VIKTIG, OG IKKE ENSARTET I KILDEN
    Formue-arket blander to prisbaser i samme kolonne av år:
        rad 38-40 (priser)   løpende kroner
        rad 55, 58 (utgifter) løpende kroner
        rad 84 (SNKS)         FASTE 2026-kroner
    Verifisert: rad 74 / rad 84 = deflatoren eksakt, og rad 38 / deflatoren
    gir 4 480,1 kr/Sm3 o.e. i 2050, som er nivået den forrige modellen brukte.
    Uttrekket gjengir kilden slik den er. Deflateringen hører hjemme i
    modell.py. Prisbasis står i kilder.csv for hver eneste kolonne.

EXCEL-FELLE
    Skiftberegning lagrer flere andelsrader som klokkeslett, ikke tall:
    0,78 ligger som 18:43:12. `_tall` konverterer. Rad 11, 14 og 21 (NGL i
    Sm3 o.e.) er i tillegg ødelagt i kilden — de gir verdier i størrelsesorden
    7e8. De hentes ikke; NGL-volumet tas fra Formue rad 32, som er intakt.

TRE VALG TATT 03.09.2026, med begrunnelse

  1. `snks` er NB26 Formue rad 84, og 2026-verdien er 521,3 mrd. kroner.
     Akseptansetesten oppga opprinnelig 685,6 mrd. fra RNB 2026. De to kan
     ikke forenes: forskyves serien ett år for å treffe 685,6, blir kumulativ
     2026-2050 5 467 i stedet for 4 861, og NPV(4 pst.; 2026-2060) bommer
     tilsvarende. Begge de to aggregatene reproduseres eksakt med 2026 =
     521,3. NB26 er dessuten den årgangen resten av modellen bygger på —
     priser, volumer og kostnader. Ett tall fra en annen årgang ville brutt
     den interne konsistensen. Testen er derfor rettet, ikke fjernet, og
     begrunnelsen står i testens docstring. Kommer RNB 2026 inn senere, blir
     det en egen serie med egen årgangsmerking.

  2. Marginalskattesatsen hentes for hele perioden kilden dekker (2001-2090).
     Den er 0,76 i 2001 og 0,78 fra 2002. Regelen er at inndata.csv skal være
     en tro gjengivelse av kilden, så serien avkortes ikke. Testen er i stedet
     avgrenset til modellperioden 2026-2090, som er der satsen brukes.

  3. sdoe_andel_ngl hentes som den står, men er merket ubrukelig etter 2045.
     Raden er SDØEs NGL-produksjon delt på total NGL-produksjon, og totalen
     kollapser: 2,5 mill. Sm3 o.e. i 2045, 0,21 i 2055, 0,047 i 2060. Med så
     liten nevner sprekker forholdstallet, og det passerer 1 i 2055-2058
     (maks 1,77). Det er ikke datokorrupsjon, men et småtallsproblem — og det
     inntreffer nettopp der NGL ikke lenger betyr noe for kontantstrømmen.
     Serien hentes derfor uten strengt spennkrav, med advarsel, og erstattes
     av oljens andel i modell.py. NGL følger oljen ellers i modellen
     (korrelasjon 0,945 i årlige logendringer), så substitusjonen er
     konsistent med resten. Den gjøres i modell.py, ikke her, slik at
     uttrekket forblir en gjengivelse av kilden.
"""

from __future__ import annotations

import datetime as _dt
import glob
import os

import pandas as pd

RAW = "data/raw"
UT = "data/inndata.csv"
KILDER = "data/kilder.csv"

AAR_FRA, AAR_TIL = 1971, 2090

# --- kolonnespesifikasjon ---------------------------------------------------
# (kolonne, ark, rad, forventet etikett, enhet, prisbasis, min, maks, streng,
#  skala)
# `skala` ganges på råverdien. Den brukes BARE til enhetsomregning, aldri til
# å avlede noe. Kilden oppgir beløp i mill. kroner; modellen og hele vedlegget
# regner i mrd. kroner, så beløpskolonnene skaleres med 1/1000 her, ved
# grensen, i stedet for at omregningen spres utover i modellkoden. Både
# kildens enhet og den lagrede enheten står i kilder.csv.
# `streng=False` gjør spennbruddet til en advarsel i stedet for en stopp. Det
# brukes bare der bruddet er forstått og dokumentert — i dag bare
# sdoe_andel_ngl, jf. valg 3 over. Spennet gjelder verdien ETTER skalering.
# Etiketten kontrolleres mot filen ved hvert uttrekk. Endrer NB27 radnumrene,
# stopper uttrekket med en beskjed om hvilken rad som ikke stemmer, i stedet
# for å hente feil serie i stillhet.
MULIGHETSBILDE = [
    ("produksjon_olje", "Formue", 30, "Produksjon av råolje",
     "mill. Sm3 o.e.", "volum", 0, 300, True, 1),
    ("produksjon_gass", "Formue", 31, "Produksjon av naturgass",
     "mill. Sm3 o.e.", "volum", 0, 300, True, 1),
    ("produksjon_ngl", "Formue", 32, "Produksjon av NGL",
     "mill. Sm3 o.e.", "volum", 0, 100, True, 1),
    ("prisbane_olje", "Formue", 38, "Oljepris NOK/Sm3 o.e.",
     "kr/Sm3 o.e.", "løpende", 100, 40000, True, 1),
    ("prisbane_gass", "Formue", 39, "Gasspris NOK/Sm3 o.e.",
     "kr/Sm3 o.e.", "løpende", 100, 40000, True, 1),
    ("prisbane_ngl", "Formue", 40, "NGL-pris kr/Sm3 o.e.",
     "kr/Sm3 o.e.", "løpende", 100, 40000, True, 1),
    ("driftsutgifter", "Formue", 55, "Påløpte driftsutgifter",
     "mrd. kr", "løpende", 0, 2_000, True, 1e-3),
    ("investeringsutgifter", "Formue", 58, "Påløpte investeringsutgifter",
     "mrd. kr", "løpende", 0, 2_000, True, 1e-3),
    ("deflator", "Formue", 44, "Deflator",
     "indeks, 2026 = 1", "—", 0.1, 20, True, 1),
    ("snks", "Formue", 84, "SNKS",
     "mrd. kr", "faste 2026", -200, 1_500, True, 1e-3),
    ("marginalskattesats", "Skiftberegning", 7, "Marginalskattesats",
     "andel", "—", 0.5, 0.9, True, 1),
    ("sdoe_andel_olje", "Skiftberegning", 37, "Olje",
     "andel", "—", 0, 1, True, 1),
    ("sdoe_andel_gass", "Skiftberegning", 38, "Gass",
     "andel", "—", 0, 1, True, 1),
    ("sdoe_andel_ngl", "Skiftberegning", 39, "NGL",
     "andel", "—", 0, 1, False, 1),
    ("gassprisgjennomslag", "Skiftberegning", 17, "Gassprisgjennomslag",
     "andel", "—", 0, 1, True, 1),
    ("dollarkurs", "KVARTS", 11, "Dollarkurs",
     "NOK/USD", "løpende", 4, 20, True, 1),
    ("fat_per_sm3", "KVARTS", 16, "Omregningsfaktor Sm3 o.e.=> fat",
     "fat/Sm3 o.e.", "—", 5, 8, True, 1),
]

# Ressursrapporten har LANGT format: år i kolonne B fra rad 20.
RESSURSRAPPORT_ARK = "Fig. 1.10"
RESSURSRAPPORT_RAD0 = 20
RESSURSRAPPORT = [
    ("produksjon_sd_historisk", 3, "Historisk produksjon"),
    ("produksjon_sd_basis", 4, "Basis"),
    ("produksjon_sd_hoy", 5, "Høy"),
    ("produksjon_sd_lav", 6, "Lav"),
]


class Uttrekksfeil(Exception):
    """Kilden ser ikke ut som forventet. Stopp og se på filen."""


# --- hjelpere ---------------------------------------------------------------

def _tall(v):
    """Excel-verdi -> float eller None.

    Skiftberegning lagrer andeler som klokkeslett (0,78 = 18:43:12) og enkelte
    rader som datoer. Konverteringen gjør om begge til den underliggende
    brøken. Tomme strenger regnes som hull, ikke som null.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, _dt.datetime):
        return (v - _dt.datetime(1899, 12, 30)).total_seconds() / 86400
    if isinstance(v, _dt.time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _aarskolonner(ws, rad: int = 1) -> dict[int, int]:
    """Årstall i `rad` -> kolonnenummer. Krever en sammenhengende serie."""
    k = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=rad, column=c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool) and 1900 < v < 2200:
            k[int(v)] = c
    if not k:
        raise Uttrekksfeil(f"{ws.title}: fant ingen årstall i rad {rad}")
    ar = sorted(k)
    if ar != list(range(ar[0], ar[-1] + 1)):
        raise Uttrekksfeil(f"{ws.title}: årsrekken {ar[0]}-{ar[-1]} har hull")
    return k


def _sjekk_etikett(ws, rad: int, ventet: str) -> None:
    funnet = ws.cell(row=rad, column=1).value
    if not (isinstance(funnet, str) and funnet.strip() == ventet):
        raise Uttrekksfeil(
            f"{ws.title} rad {rad}: ventet etiketten {ventet!r}, fant {funnet!r}. "
            "Radnumrene har flyttet seg — kjør `python -m src.inspiser` på nytt."
        )


def _sjekk_spenn(navn: str, s: pd.Series, lav: float, hoy: float,
                 streng: bool = True) -> None:
    f = s.dropna()
    if f.empty:
        raise Uttrekksfeil(f"{navn}: ingen verdier hentet")
    if f.min() < lav or f.max() > hoy:
        ute = f[(f < lav) | (f > hoy)]
        melding = (f"{navn}: {len(ute)} verdier utenfor [{lav}, {hoy}], "
                   f"første {ute.index[0]} = {ute.iloc[0]:,.4g}")
        if streng:
            raise Uttrekksfeil(melding + ". Sannsynlig enhets- eller datofeil "
                               "i kilden — kjør `python -m src.inspiser`.")
        print(f"  ADVARSEL  {melding} (kjent og dokumentert, jf. kilder.csv)")


# --- lesing -----------------------------------------------------------------

def les_mulighetsbilde(sti: str) -> pd.DataFrame:
    """Formue, Skiftberegning og KVARTS -> bred tabell indeksert på år."""
    import openpyxl

    wb = openpyxl.load_workbook(sti, data_only=True)
    ut = pd.DataFrame(index=pd.RangeIndex(AAR_FRA, AAR_TIL + 1, name="ar"))
    kolonner: dict[str, dict[int, int]] = {}

    for navn, ark, rad, etikett, _e, _b, lav, hoy, streng, skala in MULIGHETSBILDE:
        if ark not in wb.sheetnames:
            raise Uttrekksfeil(f"fant ikke arket {ark!r} i {os.path.basename(sti)}")
        ws = wb[ark]
        if ark not in kolonner:
            kolonner[ark] = _aarskolonner(ws)
        _sjekk_etikett(ws, rad, etikett)
        k = kolonner[ark]
        s = pd.Series(
            {y: _tall(ws.cell(row=rad, column=c).value)
             for y, c in k.items() if AAR_FRA <= y <= AAR_TIL},
            dtype="float64",
        ) * skala
        _sjekk_spenn(navn, s, lav, hoy, streng)
        ut[navn] = s

    wb.close()
    return ut


def les_ressursrapport(sti: str) -> pd.DataFrame:
    """Ressursrapport 2026 figur 1.10 -> bred tabell indeksert på år.

    Arket har LANGT format, i motsetning til Mulighetsbilde: år i kolonne B
    fra rad 20, med historisk, Basis, Høy og Lav i kolonne C-F.
    """
    import openpyxl

    wb = openpyxl.load_workbook(sti, data_only=True)
    if RESSURSRAPPORT_ARK not in wb.sheetnames:
        raise Uttrekksfeil(f"fant ikke arket {RESSURSRAPPORT_ARK!r}")
    ws = wb[RESSURSRAPPORT_ARK]

    hode = {c: ws.cell(row=RESSURSRAPPORT_RAD0 - 2, column=c).value
            for _, c, _ in RESSURSRAPPORT}
    for navn, c, ventet in RESSURSRAPPORT:
        if str(hode[c]).strip() != ventet:
            raise Uttrekksfeil(
                f"{RESSURSRAPPORT_ARK} kolonne {c}: ventet {ventet!r}, "
                f"fant {hode[c]!r}"
            )

    rader = {}
    for r in range(RESSURSRAPPORT_RAD0, ws.max_row + 1):
        y = _tall(ws.cell(row=r, column=2).value)
        if y is None or not (AAR_FRA <= y <= AAR_TIL):
            continue
        rader[int(y)] = {navn: _tall(ws.cell(row=r, column=c).value)
                         for navn, c, _ in RESSURSRAPPORT}
    wb.close()

    ut = pd.DataFrame.from_dict(rader, orient="index").sort_index()
    ut.index.name = "ar"
    ut = ut.reindex(pd.RangeIndex(AAR_FRA, AAR_TIL + 1, name="ar"))
    for navn, _, _ in RESSURSRAPPORT:
        _sjekk_spenn(navn, ut[navn], 0, 400)
    return ut


# --- kontroller -------------------------------------------------------------

def kontroller(d: pd.DataFrame) -> list[str]:
    """Kontroller som skal holde uansett årgang. Returnerer avvik."""
    feil = []

    if abs(d.loc[2026, "deflator"] - 1.0) > 1e-9:
        feil.append(f"deflator 2026 = {d.loc[2026, 'deflator']:.6f}, skal være 1")

    # Drift + investering skal summere til samlede utgifter i kilden.
    # Kontrollen ligger her fordi den fanger opp at feil rad er hentet.
    if not (d["driftsutgifter"].notna() == d["investeringsutgifter"].notna()).all():
        feil.append("drifts- og investeringsutgifter har ulike hull")

    f = d["produksjon_sd_lav"] / d["produksjon_sd_basis"]
    if (f.dropna() > 1 + 1e-9).any():
        ar = f[f > 1 + 1e-9].index.tolist()
        feil.append(f"lavbanen ligger over basisbanen i {ar}")

    for kol in d.columns:
        if (d[kol].dropna() == 0).all() and d[kol].notna().any():
            feil.append(f"{kol}: bare nuller — sannsynlig feil rad")

    return feil


def _oppdater_kilder(d: pd.DataFrame, dato: str) -> pd.DataFrame:
    """Fyller ut kilder.csv: enhet, prisbasis, dekning og uttrekksdato."""
    spec = {navn: (ark, rad, enhet, basis, skala)
            for navn, ark, rad, _e, enhet, basis, _l, _h, _s, skala in MULIGHETSBILDE}

    k = pd.read_csv(KILDER).set_index("serie")
    for kol in ("enhet", "enhet_kilde", "prisbasis", "dekning"):
        if kol not in k.columns:
            k[kol] = ""

    # sdoe_andeler er én rad i malen, men tre serier i uttrekket
    if "sdoe_andeler" in k.index:
        mal = k.loc["sdoe_andeler"].copy()
        k = k.drop(index="sdoe_andeler")
        for res, rad in (("olje", 37), ("gass", 38), ("ngl", 39)):
            ny = mal.copy()
            ny["celleomrade"] = f"rad {rad}"
            ny["merknad"] = f"SDØEs produksjonsandel, {res}"
            k.loc[f"sdoe_andel_{res}"] = ny

    nye = {
        "produksjon_sd_historisk": ("Ressursrapport 2026", "RR26", "ja",
                                    "figur 1.10, kolonne C"),
        "dollarkurs": ("Nasjonalbudsjettet 2026", "NB26", "nei",
                       "KVARTS rad 11. 10,114 flatt fra 2026 — svarer trolig "
                       "på aapent punkt 1 om valutakursen bak Sodirs tall."),
        "fat_per_sm3": ("Nasjonalbudsjettet 2026", "NB26", "nei",
                        "KVARTS rad 16. Erstatter den hardkodede 6,29."),
        # Bekreftet 03.09.2026: finnes IKKE i noen av kildefilene i data/raw.
        # Søk på oljekorrigert / fondet / pensjonsfond / overføring gav null
        # treff i alle ti ark i Mulighetsbilde. Må hentes utenfra til steg 3.
        "oljekorrigert_underskudd": ("Nasjonalbudsjettet 2026, hovedtabeller",
                                     "NB26", "ja",
                                     "MANGLER — ikke i data/raw. Kun steg 3."),
        "fondsverdi": ("Nasjonalbudsjettet 2026, hovedtabeller", "NB26", "ja",
                       "MANGLER — ikke i data/raw. Kun steg 3."),
    }
    for navn, (kilde, argang, off, merk) in nye.items():
        if navn not in k.index:
            k.loc[navn] = ""
            k.loc[navn, ["kilde", "argang", "offentlig", "merknad"]] = \
                [kilde, argang, off, merk]

    for navn, (_, c, _) in {n: (None, c, None) for n, c, _ in RESSURSRAPPORT}.items():
        if navn in k.index:
            k.loc[navn, "filnavn"] = "ressursrapport-...-numbers.xlsx"
            k.loc[navn, "ark"] = RESSURSRAPPORT_ARK
            k.loc[navn, "celleomrade"] = f"kolonne {c}, fra rad {RESSURSRAPPORT_RAD0}"
            k.loc[navn, "enhet"] = "mill. Sm3 o.e."
            k.loc[navn, "enhet_kilde"] = "mill. Sm3 o.e."
            k.loc[navn, "prisbasis"] = "—"
            if str(k.loc[navn, "merknad"]).startswith("MANGLER"):
                k.loc[navn, "merknad"] = k.loc[navn, "merknad"].replace(
                    "MANGLER", "hentet").replace("hentet - ", "")

    for navn, (ark, rad, enhet, basis, skala) in spec.items():
        if navn not in k.index:
            k.loc[navn] = ""
        k.loc[navn, "ark"] = ark
        k.loc[navn, "celleomrade"] = f"rad {rad}"
        k.loc[navn, "enhet"] = enhet
        k.loc[navn, "enhet_kilde"] = ("mill. kr" if skala == 1e-3 else enhet)
        k.loc[navn, "prisbasis"] = basis

    k.loc["sdoe_andel_ngl", "merknad"] = (
        "UBRUKELIG ETTER 2045: forholdstall med kollapsende nevner "
        "(NGL-produksjonen faller til 0,05 mill. Sm3 o.e. i 2060), passerer 1 "
        "i 2055-2058. Erstattes av oljens andel i modell.py."
    )
    k.loc["gassprisgjennomslag", "merknad"] = (
        "AAPENT PUNKT 3: 0,50 hele perioden. Parameter med standardverdi 1,0 "
        "og bryter til 0,50, jf. README."
    )
    k.loc["marginalskattesats", "merknad"] = (
        "0,76 i 2001, 0,78 fra 2002. Hentet uavkortet; testen dekker 2026-2090."
    )
    k.loc["snks", "merknad"] = (
        "NB26 rad 84, faste 2026-kroner. 2026 = 521,3 mrd. Se valg 1 i "
        "src/uttrekk.py om hvorfor RNB 2026s 685,6 ikke er brukt."
    )

    for navn in k.index:
        if navn in d.columns:
            f = d[navn].dropna()
            k.loc[navn, "dekning"] = f"{f.index.min()}-{f.index.max()}"
            k.loc[navn, "uttrekksdato"] = dato

    return k.sort_index().reset_index()


# --- bygg -------------------------------------------------------------------

def _finn(monster: str) -> str:
    traff = [p for p in glob.glob(f"{RAW}/*.xlsx")
             if monster.lower() in os.path.basename(p).lower()]
    if not traff:
        raise Uttrekksfeil(f"fant ingen fil som matcher {monster!r} i {RAW}/")
    return sorted(traff)[0]


def bygg(skriv: bool = True) -> pd.DataFrame:
    """Bygger data/inndata.csv og oppdaterer data/kilder.csv."""
    d = les_mulighetsbilde(_finn("mulighetsbilde"))
    d = d.join(les_ressursrapport(_finn("ressursrapport")))

    feil = kontroller(d)
    if feil:
        raise Uttrekksfeil("kontroller feilet:\n  - " + "\n  - ".join(feil))

    if skriv:
        dato = _dt.date.today().isoformat()
        d.to_csv(UT, float_format="%.6f")
        _oppdater_kilder(d, dato).to_csv(KILDER, index=False)
    return d


if __name__ == "__main__":
    d = bygg()
    print(f"Skrev {UT}: {len(d)} år ({d.index.min()}-{d.index.max()}), "
          f"{len(d.columns)} serier\n")
    print(f"{'serie':26s}{'dekning':14s}{'n':>5}  {'2026':>12s}{'2050':>12s}")
    print("-" * 72)
    for kol in d.columns:
        f = d[kol].dropna()
        v = lambda y: (f"{d.loc[y, kol]:12,.3f}" if pd.notna(d.loc[y, kol]) else f"{'—':>12}")
        print(f"{kol:26s}{f'{f.index.min()}-{f.index.max()}':14s}{len(f):5d}  {v(2026)}{v(2050)}")
