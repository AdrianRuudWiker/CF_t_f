# -*- coding: utf-8 -*-
"""
Regenererer viftefiguren for den REFORMULERTE usikkerhetsmodellen, konsistent
med arket «Reformulert vifte» og den skjulte motoren «MC-motor-R».

Leser de faste trekkene (w, z1, z2) og alle forutsetninger direkte ut av
Kontantstromsmodell_petroleum.xlsx, og regner SNCF med samme uttrykk som
Excel-motoren, slik at figuren viser nøyaktig de samme tallene som arbeidsboken.

Begge forankringer er med, etter brukerens beslutning: persentilbåndene og
medianen er MEDIANFORANKRET (hovedsporet, der P50 = NB26), og den
forventningsforankrede medianen ligger inne som egen linje, slik at leseren ser
hva forankringsvalget faktisk gjør. Fire linjer er maksgrensen i profilen; her
er det tre pluss båndene.

Figurer:
  reformulert_vifte.svg      - årlig SNCF, tetthetsvifte P5-P95, begge medianer
  reformulert_akkumulert.svg - akkumulert SNCF, persentilbånd, begge medianer

Kjøring:  python3 lag_figur_reformulert.py
Krever:   numpy, openpyxl, matplotlib
"""
import sys

import numpy as np
from openpyxl import load_workbook

sys.path.insert(0, "/root/.claude/skills/synced/"
                "5030c864-d034-405a-808f-f52ea4fe255a_9e3f4a54-160c-4203-91ed-822663bb2dc7/"
                "fin-designprofil/scripts")
from fin_chart_style import (apply_matplotlib_style, FIN_DARKBLUE, FIN_BLUE,
                             FIN_LIGHTBLUE, FIN_RED)
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

import build_reformulert as br

apply_matplotlib_style()
YEARS = np.arange(2026, 2051)


def nf(x, dec=0):
    """Norsk tallformat: mellomrom som tusenskille, komma som desimaltegn."""
    return f"{x:,.{dec}f}".replace(",", " ").replace(".", ",")


def simuler():
    """SNCF for begge forankringer på arkets egne faste trekk."""
    wb = load_workbook(br.FIL, data_only=True)
    fu = wb["Forutsetninger"]
    col = lambda L: np.array([fu[f"{L}{18 + i}"].value for i in range(25)])
    volO, volG, volN = col("B"), col("C"), col("D")
    totH, totL = col("F"), col("G")
    pO, pG, pN, cost, snks = col("J"), col("K"), col("L"), col("M"), col("N")
    tot = volO + volG + volN
    fh, fl = totH / tot, totL / tot
    nks = volO * pO + volG * pG + volN * pN - cost
    andel = snks / nks
    basis = andel * nks / 1000.0                       # NB26-basis per år (mrd.)

    # Sigma avledes av persentilforholdene, som i arket.
    from statistics import NormalDist
    z90 = NormalDist().inv_cdf(0.9)
    sig_o = np.log(fu["B4"].value) / z90
    sig_g = np.log(fu["B6"].value) / z90
    rho = fu[f"B{br.R_RHO}"].value

    # Trekkene ligger i motoren; les dem derfra så figuren og arket er identiske.
    m = wb[br.MOTOR]
    n = fu[f"B{br.R_N}"].value
    w = np.array([m.cell(br.DATA0 + r, br.C_W).value for r in range(n)])
    z1 = np.array([m.cell(br.DATA0 + r, br.C_Z1).value for r in range(n)])
    z2 = np.array([m.cell(br.DATA0 + r, br.C_Z2).value for r in range(n)])

    fo = np.exp(sig_o * z1)[:, None]
    fg = np.exp(sig_g * (rho * z1 + np.sqrt(1 - rho ** 2) * z2))[:, None]
    volfac = np.where(w[:, None] >= 0, 1 + w[:, None] * (fh - 1),
                      1 + w[:, None] * (1 - fl)) / (1 + (fh + fl - 2) / 6)
    ut = {}
    for navn, ko, kg in (("med", 1.0, 1.0),
                         ("for", np.exp(-sig_o ** 2 / 2), np.exp(-sig_g ** 2 / 2))):
        netto = volfac * (volO * pO * fo * ko + volN * pN * fo * ko
                          + volG * pG * fg * kg - cost)
        ut[navn] = andel * np.maximum(netto, 0.0) / 1000.0
    return ut, basis


def _fin_akser(ax):
    """Y-akse på begge sider, tickmarks innover, ingen x-tickmarks."""
    ax.tick_params(axis="y", which="both", direction="in", right=True, left=True,
                   labelright=False, length=3)
    ax.tick_params(axis="x", length=0)
    for s in ax.spines.values():
        s.set_linewidth(0.5)
        s.set_color("#000000")
    ax.spines["top"].set_visible(False)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: nf(v)))


def _vifte(ax, data, farge=FIN_BLUE):
    """Gradert tetthetsvifte, kuttet ved P5-P95."""
    for lo in np.arange(5, 50, 5):
        ax.fill_between(YEARS, np.percentile(data, lo, axis=0),
                        np.percentile(data, 100 - lo, axis=0), color=farge,
                        alpha=0.10 + 0.55 * (lo / 45.0), linewidth=0, zorder=1)


def _median(ax, y, etikett):
    """Medianbanen: mørk kontur med hvit kjerne, som i de øvrige figurene.

    Etiketten legges på den MØRKE konturen, ikke på den hvite kjernen — en hvit
    linje blir usynlig i tegnforklaringen.
    """
    ax.plot(YEARS, y, color=FIN_DARKBLUE, lw=3.0, zorder=5, label=etikett)
    ax.plot(YEARS, y, color="white", lw=1.4, zorder=6)


def figur_vifte(ut, basis):
    fig, ax = plt.subplots(figsize=(8.9, 4.72))
    _vifte(ax, ut["med"])
    ax.plot(YEARS, basis, color=FIN_RED, ls="--", lw=1.5, zorder=4,
            label="Basisbane (NB26 / IEA WEO APS)")
    ax.plot(YEARS, np.median(ut["for"], axis=0), color=FIN_LIGHTBLUE, ls=":",
            lw=1.5, zorder=4, label="Median, forventningsforankret")
    _median(ax, np.median(ut["med"], axis=0), "Median, medianforankret (P50)")
    ax.axhline(0, color="#000000", lw=0.5, zorder=2)
    ax.set_xlim(2026, 2050)
    ax.margins(x=0)
    ax.set_xticks([2030, 2035, 2040, 2045, 2050])
    _fin_akser(ax)
    ax.annotate("Mrd. 2026-kroner", xy=(0, 1.02), xycoords="axes fraction",
                fontsize=9, color="#000000")
    ax.legend(loc="upper right", fontsize=9)
    fig.tight_layout()
    fig.savefig("reformulert_vifte.svg")
    plt.close(fig)


def figur_akkumulert(ut, basis):
    fig, ax = plt.subplots(figsize=(8.9, 4.72))
    kum = {k: np.cumsum(v, axis=1) for k, v in ut.items()}
    _vifte(ax, kum["med"])
    ax.plot(YEARS, np.cumsum(basis), color=FIN_RED, ls="--", lw=1.5, zorder=4,
            label="Basisbane (NB26 / IEA WEO APS)")
    ax.plot(YEARS, np.median(kum["for"], axis=0), color=FIN_LIGHTBLUE, ls=":",
            lw=1.5, zorder=4, label="Median, forventningsforankret")
    _median(ax, np.median(kum["med"], axis=0), "Median, medianforankret (P50)")
    ax.set_xlim(2026, 2050)
    ax.margins(x=0)
    ax.set_xticks([2030, 2035, 2040, 2045, 2050])
    _fin_akser(ax)
    ax.annotate("Mrd. 2026-kroner, akkumulert", xy=(0, 1.02),
                xycoords="axes fraction", fontsize=9, color="#000000")
    ax.legend(loc="upper left", fontsize=9)
    fig.tight_layout()
    fig.savefig("reformulert_akkumulert.svg")
    plt.close(fig)


if __name__ == "__main__":
    ut, basis = simuler()
    figur_vifte(ut, basis)
    figur_akkumulert(ut, basis)
    kum = ut["med"].sum(axis=1)
    print(f"Skrev reformulert_vifte.svg og reformulert_akkumulert.svg")
    print(f"Kontroll: kumulativ P50 {np.percentile(kum, 50):.0f} mrd. mot basis "
          f"{basis.sum():.0f} mrd. ({100 * (np.percentile(kum, 50) / basis.sum() - 1):+.1f} pst.)")
