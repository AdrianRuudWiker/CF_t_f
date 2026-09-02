# -*- coding: utf-8 -*-
"""
Bygger den REFORMULERTE usikkerhetsmodellen inn i
Kontantstromsmodell_petroleum.xlsx, Excel-native.

Beslutninger dette bygget realiserer (avklart med brukeren 02.09.2026):
- Persentilene ER scenariene: P90 = høy prognosert CF, P50 = median, P10 = lav.
  Ingen 3x3 av vilkårlige multiplikatorer.
- BEGGE forankringer vises: medianforankring (P50 = NB26) som hovedspor og
  forventningsforankring (E = NB26) som følsomhet, side om side i samme ark,
  drevet av SAMME trekk. Forskjellen er én multiplikativ Jensen-korreksjon
  EXP(-sigma^2/2) på prisfaktoren, så motoren trenger ikke to sett trekk.
- Kalibreringsvei (c): sigma avledes historisk av Forutsetninger!B4:B7
  (sigma = LN(persentilforhold)/NORMSINV(0,9), eksakt for en medianforankret
  lognormal), og IEA WEO NZE legges inn som et NAVNGITT SIDESCENARIO utenfor
  viften — ikke som en persentil. Viften viser markedsrisiko kalibrert på
  historikk; NZE-linjen viser politikkrisiko historikken ikke inneholder.
- Balansepris-gulv: feltnetto gulves ved 0, styrt av en bryter i arket.

Ikke-destruktivt: de eksisterende arkene (OU-motoren, 3x3-en, Miksfølsomhet)
står urørt. Bygget legger til to ark:
- "MC-motor-R" (skjult): trekk + levende formler, begge forankringer.
- "Reformulert vifte" (synlig): persentiltabell, oppsummering, NZE-sidescenario
  og parameteravlesning — alt PERCENTILE/AVERAGE-formler mot motoren.

NZE-sidescenarioet krever tall fra IEA WEO Annex A som IKKE kunne hentes
(egress-proxyen blokkerer iea.org). Cellene er lagt inn som blå input og står
tomme; scenarioet gir #N/A til de er fylt ut, i stedet for å vise et tall som
ikke har kilde.

Kjøring:  python3 build_reformulert.py
Krever:   numpy, openpyxl
"""
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as CL

FIL = "Kontantstromsmodell_petroleum.xlsx"
MOTOR, VIFTE, UTV = "MC-motor-R", "Reformulert vifte", "Utvidelse 2060"

YEARS = list(range(2026, 2051))     # SD-dekket, kildebelagt
YEARS_EXT = list(range(2051, 2061))  # ekstrapolert, se arket «Utvidelse 2060»
YEARS_ALL = YEARS + YEARS_EXT
NY, NYE, NYA = len(YEARS), len(YEARS_EXT), len(YEARS_ALL)
FU0 = 18                    # Forutsetninger: 2026 på rad 18
EXT0 = 19                   # Utvidelse 2060: 2051 på rad 19
N = 2000                    # simuleringer
SEED = 2026

DATA0 = 10                  # motor: første simulering
C_SIM, C_W, C_Z1, C_Z2 = 1, 2, 3, 4
C_FO, C_FG = 5, 6           # prisfaktorer, beregnet én gang per simulering
C_MED, C_FOR = 8, 44        # SNCF-blokker (35 kolonner hver: 2026-2060)
# Aggregater per forankring: KUM50, NPV3_50, NPV2_50, NPV4_50, KUM60,
# NPV4_60, NPV3_60 — 2050-horisonten først, så 2060-horisonten.
A_MED, A_FOR = 80, 89
AGG = ("KUM50", "NPV3_50", "NPV2_50", "NPV4_50", "KUM60", "NPV4_60", "NPV3_60")

# Forutsetninger: nytt parameterblokk under kildenoten (rad 44)
P0 = 46
R_SIGO, R_SIGG, R_RHO, R_GULV = P0 + 1, P0 + 2, P0 + 3, P0 + 4
R_JO, R_JG, R_N, R_SEED = P0 + 5, P0 + 6, P0 + 7, P0 + 8
R_BBL, R_FX, R_BASO, R_BASG, R_KRG = P0 + 9, P0 + 10, P0 + 11, P0 + 12, P0 + 13
R_NZEO, R_NZEG, R_STPO, R_STPG = P0 + 14, P0 + 15, P0 + 16, P0 + 17
R_PNZEO, R_PNZEG, R_PSTPO, R_PSTPG = P0 + 18, P0 + 19, P0 + 20, P0 + 21
R_NZEOK = P0 + 22           # er NZE-scenarioet komplett?

C_NZEO, C_NZEG = 17, 18     # årstabell: Q, R = NZE-prisbaner (input)

# Utvidelse 2060: parameterrader på det nye arket
E_VINDU, E_RJUST = 4, 5
E_ROLJE, E_RGASS, E_RNGL = 6, 7, 8
E_RTOTH, E_RTOTL, E_RKOST = 9, 10, 11
E_ANDEL, E_PM, E_DEFL = 12, 13, 14
E_NB26, E_HALE = 15, 16
E_BAS0 = 5                  # samlet basisbane 2026-2060 i kolonne R/S, rad 5-39

BLA = Font(color="FF0000FF")
HDR = Font(color="FFFFFFFF", bold=True)
HDRF = PatternFill("solid", fgColor="FF181C62")
FET = Font(bold=True)
GRA = Font(italic=True, color="FF595959")
F2, F0 = "#,##0.00", "#,##0"


def trekk(n=N, seed=SEED):
    """Faste, persistente regimetrekk — ett per simulering.

    Samme rekkefølge som mc_reformulert.simuler (z før w), slik at Excel og
    Python-referansen kan sammenlignes på identiske trekk.
    """
    rng = np.random.default_rng(seed)
    z = rng.standard_normal((n, 2))
    w = rng.triangular(-1, 0, 1, n)
    return w, z[:, 0], z[:, 1]


def _fu(ref):
    return f"Forutsetninger!{ref}"


def _utv(ref):
    return f"'{UTV}'!{ref}"


def _aarskilde(i):
    """Returnerer en funksjon kol -> cellereferanse for årsindeks i.

    Årene 2026-2050 ligger i Forutsetninger, 2051-2060 i «Utvidelse 2060».
    Det nye arket bruker SAMME kolonnebokstaver, så kallerne slipper å vite
    hvilken kilde året kommer fra.
    """
    if i < NY:
        rad = FU0 + i
        return lambda kol: _fu(f"{kol}{rad}")
    rad = EXT0 + (i - NY)
    return lambda kol: _utv(f"{kol}{rad}")


def _parametre(wb):
    """Nytt parameterblokk i Forutsetninger. Idempotent — skrives alltid likt."""
    fu = wb["Forutsetninger"]
    fu.cell(P0, 1, "Parametre — reformulert usikkerhetsmodell "
                   "(persentiler som scenarier)").font = FET

    rader = [
        (R_SIGO, "sigma_olje (persistent prisregime)",
         "=LN(B4)/NORMSINV(0.9)",
         "Avledet: for en medianforankret lognormal er sigma = LN(P90/P50)/z90. "
         "Historisk kalibrering, jf. B4/B5."),
        (R_SIGG, "sigma_gass (persistent prisregime)",
         "=LN(B6)/NORMSINV(0.9)",
         "Avledet på samme måte fra B6/B7."),
        (R_RHO, "korrelasjon olje/gass (regime)", 0.60,
         "Input. Regimetrekkene er korrelerte, men gass har egen dynamikk."),
        (R_GULV, "Balansepris-gulv (1 = på, 0 = av)", 1,
         "Tilbudsrespons: feltnetto gulves ved 0, ingen tapsproduksjon. "
         "Forankret i balanseprisene, deck slide 14 (~20-45 USD/fat)."),
        (R_JO, "Jensen-korreksjon olje", f"=EXP(-(B{R_SIGO}^2)/2)",
         "1/E[EXP(sigma*z)]. Brukes KUN i forventningsforankringen. "
         "Parentesen om kvadratet er nødvendig: Excel binder unær minus "
         "sterkere enn potens, så -B^2 ville blitt (-B)^2 = +B^2."),
        (R_JG, "Jensen-korreksjon gass", f"=EXP(-(B{R_SIGG}^2)/2)",
         "Samme for gass."),
        (R_N, "Antall simuleringer", N, "Faste trekk i MC-motor-R."),
        (R_SEED, "Frø (seed)", SEED, "Reproduserbart."),
        (R_BBL, "Fat per Sm3 o.e.", 6.2898, "Input, enhetskonvertering."),
        (R_FX, "NOK per USD", 10.5, "Input, enhetskonvertering."),
        (R_BASO, "Impliert basis oljepris USD/fat (2050)",
         f"=J42/(B{R_BBL}*B{R_FX})",
         "Kontroll mot deck slide 5 (70 USD/fat fra 2035). Modellen ligger "
         "marginalt lavere; avviket er dokumentert."),
        (R_BASG, "Basis gasspris USD/MMBtu (fra 2040)", 5.7,
         "Deck slide 5. Input, brukes til enhetskonvertering for gass."),
        (R_KRG, "kr/Sm3 o.e. per USD/MMBtu", f"=K42/B{R_BASG}",
         "Avledet av basisbanen, så gasspriser i USD kan legges inn direkte."),
        (R_NZEO, "IEA WEO NZE oljepris USD/fat — FYLL INN", None,
         "Annex A. Kunne ikke hentes (egress blokkert). Søketreff antydet "
         "25 USD/fat i 2050 for WEO 2025, men tallet er UVERIFISERT og er "
         "derfor ikke lagt inn."),
        (R_NZEG, "IEA WEO NZE gasspris USD/MMBtu — FYLL INN", None,
         "Annex A. Ikke funnet i det hele tatt."),
        (R_STPO, "IEA WEO STEPS oljepris USD/fat — FYLL INN", None,
         "Annex A. Kun til persentilavlesningen under, ikke til kalibrering."),
        (R_STPG, "IEA WEO STEPS gasspris USD/MMBtu — FYLL INN", None,
         "Annex A."),
        (R_PNZEO, "NZE olje faller på persentil",
         f'=IF(ISNUMBER(B{R_NZEO}),NORMSDIST(LN(B{R_NZEO}/B{R_BASO})/B{R_SIGO}),"")',
         "Leser av hvor scenarioet havner i den historisk kalibrerte viften. "
         "Viser hvor uenige kildene er — IEA kalibrerer ikke viften."),
        (R_PNZEG, "NZE gass faller på persentil",
         f'=IF(ISNUMBER(B{R_NZEG}),NORMSDIST(LN(B{R_NZEG}/B{R_BASG})/B{R_SIGG}),"")',
         ""),
        (R_PSTPO, "STEPS olje faller på persentil",
         f'=IF(ISNUMBER(B{R_STPO}),NORMSDIST(LN(B{R_STPO}/B{R_BASO})/B{R_SIGO}),"")',
         ""),
        (R_PSTPG, "STEPS gass faller på persentil",
         f'=IF(ISNUMBER(B{R_STPG}),NORMSDIST(LN(B{R_STPG}/B{R_BASG})/B{R_SIGG}),"")',
         ""),
        (R_NZEOK, "NZE-sidescenario komplett (1/0)",
         f"=IF(AND(OR(COUNT(Q{FU0}:Q42)={NY},ISNUMBER(B{R_NZEO})),"
         f"OR(COUNT(R{FU0}:R42)={NY},ISNUMBER(B{R_NZEG}))),1,0)",
         "Scenarioet gir #N/A til både olje- og gassprisen er lagt inn, "
         "enten som flatt nivå over eller som årsbane i kolonne Q/R."),
    ]
    inputrader = {R_RHO, R_GULV, R_N, R_SEED, R_BBL, R_FX, R_BASG,
                  R_NZEO, R_NZEG, R_STPO, R_STPG}
    for rad, navn, verdi, note in rader:
        fu.cell(rad, 1, navn)
        c = fu.cell(rad, 2, verdi)
        if rad in inputrader:
            c.font = BLA
        if rad in (R_PNZEO, R_PNZEG, R_PSTPO, R_PSTPG):
            c.number_format = "0.0 %"
        elif rad not in (R_GULV, R_N, R_SEED):
            c.number_format = F2
        fu.cell(rad, 3, note)

    # Årstabell: input-kolonner for NZE-prisbaner (valgfritt alternativ til
    # det flate nivået i parameterblokken).
    for col, tekst in ((C_NZEO, "NZE oljepris (USD/fat) — input, Annex A"),
                       (C_NZEG, "NZE gasspris (USD/MMBtu) — input, Annex A")):
        h = fu.cell(16, col, tekst)
        h.font, h.fill = HDR, HDRF
        for i in range(NY):
            fu.cell(FU0 + i, col).font = BLA
        fu.column_dimensions[CL(col)].width = 30


def _motor(wb, w, z1, z2, n=N):
    """Skjult motor: faste trekk + levende formler, begge forankringer."""
    if MOTOR in wb.sheetnames:
        del wb[MOTOR]
    ws = wb.create_sheet(MOTOR)
    ws.sheet_state = "hidden"

    ws.cell(1, 1, f"MC-MOTOR, REFORMULERT: {n} simuleringer, persistente "
                  f"regimetrekk, {YEARS_ALL[0]}-{YEARS_ALL[-1]}. Begge "
                  f"forankringer på samme trekk. Årene etter "
                  f"{YEARS[-1]} er ekstrapolert — se arket «{UTV}».").font = FET
    for r, lbl in zip(range(3, 9), ("A = volO*pO + volN*pN", "B = volG*pG",
                                    "C = kostnader", "andel", "fh", "fl")):
        ws.cell(r, 2, lbl)

    hdr = ["sim", "w", "z1", "z2", "fo", "fg", ""] \
        + [f"SNCFmed_{y}" for y in YEARS_ALL] + [""] \
        + [f"SNCFfor_{y}" for y in YEARS_ALL] + [""] \
        + [f"{a}med" for a in AGG] + ["", ""] \
        + [f"{a}for" for a in AGG]
    for c, h in enumerate(hdr, 1):
        if h:
            ws.cell(9, c, h).font = FET

    # Per-år parameterblokk over medianblokken; forventningsblokken peker på
    # de samme cellene, så tallgrunnlaget står ett sted. Arket «Utvidelse 2060»
    # speiler Forutsetninger kolonne for kolonne, så formlene er identiske og
    # bare kilden skifter etter 2050.
    for i in range(NYA):
        ref = _aarskilde(i)
        ws.cell(3, C_MED + i, f"={ref('B')}*{ref('J')}+{ref('D')}*{ref('L')}")
        ws.cell(4, C_MED + i, f"={ref('C')}*{ref('K')}")
        ws.cell(5, C_MED + i, f"={ref('M')}")
        ws.cell(6, C_MED + i, f"={ref('P')}")
        ws.cell(7, C_MED + i, f"={ref('H')}")
        ws.cell(8, C_MED + i, f"={ref('I')}")

    sig_o, sig_g = _fu(f"$B${R_SIGO}"), _fu(f"$B${R_SIGG}")
    rho, gulv = _fu(f"$B${R_RHO}"), _fu(f"$B${R_GULV}")
    jo, jg = _fu(f"$B${R_JO}"), _fu(f"$B${R_JG}")

    fo_c, fg_c = CL(C_FO), CL(C_FG)
    for r in range(n):
        row = DATA0 + r
        ws.cell(row, C_SIM, r + 1)
        ws.cell(row, C_W, float(w[r]))
        ws.cell(row, C_Z1, float(z1[r]))
        ws.cell(row, C_Z2, float(z2[r]))
        # Prisfaktorene beregnes ÉN gang per simulering (persistent regime),
        # slik at årsformlene blir korte. Gass korreleres med olje i formelen.
        ws.cell(row, C_FO, f"=EXP({sig_o}*$C{row})")
        ws.cell(row, C_FG, f"=EXP({sig_g}*({rho}*$C{row}"
                           f"+SQRT(1-{rho}^2)*$D{row}))")
        for base, kor_o, kor_g in ((C_MED, "", ""), (C_FOR, f"*{jo}", f"*{jg}")):
            for i in range(NYA):
                pc = CL(C_MED + i)      # parameterkolonnen ligger i medianblokken
                volfac = (f"IF($B{row}>=0,1+$B{row}*({pc}$7-1),"
                          f"1+$B{row}*(1-{pc}$8))")
                netto = (f"{volfac}/(1+({pc}$7+{pc}$8-2)/6)"
                         f"*({pc}$3*${fo_c}{row}{kor_o}"
                         f"+{pc}$4*${fg_c}{row}{kor_g}-{pc}$5)")
                # Gulvbryter uten å gjenta netto-uttrykket: nedre grense er 0
                # når gulv=1 og -1E30 (praktisk talt ingen grense) når gulv=0.
                ws.cell(row, base + i,
                        f"={pc}$6*MAX({netto},({gulv}-1)*1E+30)/1000")
        for base, agg in ((C_MED, A_MED), (C_FOR, A_FOR)):
            r50 = f"{CL(base)}{row}:{CL(base + NY - 1)}{row}"
            r60 = f"{CL(base)}{row}:{CL(base + NYA - 1)}{row}"
            # NPV() diskonterer første beløp ett år, så alle nåverdier er
            # datert 2025 — samme dato som PM-referansen «NNV i 2025».
            for off, f in enumerate((
                    f"=SUM({r50})",
                    f"=NPV({_fu('$B$8')},{r50})",
                    f"=NPV({_fu('$B$9')},{r50})",
                    f"=NPV({_fu('$B$10')},{r50})",
                    f"=SUM({r60})",
                    f"=NPV({_fu('$B$10')},{r60})",
                    f"=NPV({_fu('$B$8')},{r60})")):
                ws.cell(row, agg + off, f)
    return ws



def _utvidelse(wb, n=N):
    """Arket «Utvidelse 2060»: ekstrapolert 2051-2060 og broen til PM-tallet.

    Årstabellen speiler Forutsetninger kolonne for kolonne (B/C/D volum,
    F/G totalbaner, H/I volumfaktorer, J/K/L priser, M kostnader, O NKS,
    P statsandel), slik at MC-motoren kan bruke identiske formler og bare
    skifte kilde etter 2050. Én forskjell i retning: fram til 2050 er SNKS
    gitt av NB26 og statsandelen avledet; etter 2050 finnes ingen NB26-anslag,
    så statsandelen holdes fast og SNCF avledes av den.

    Alle oppslag bruker AVGRENSEDE områder (B17:B42), ikke hele kolonner.
    Fullkolonnereferanser er gyldig Excel, men får formeltolkere til å bygge
    en graf over en million celler.
    """
    if UTV in wb.sheetnames:
        del wb[UTV]
    ws = wb.create_sheet(UTV, 5)

    ws["A1"] = ("UTVIDELSE TIL 2060 — ekstrapolert, og sammenligning mot "
                "PM-referansen på 4 800 mrd.")
    ws["A1"].font = FET
    ws["A2"] = (
        "Sokkeldirektoratets mulighetsbilder slutter i 2050, så 2051-2060 er "
        "EKSTRAPOLERT og har ingen kilde. Volumer, totalbaner og kostnader "
        "føres videre med den geometriske årsraten observert over de siste "
        "årene av basisbanen (vinduet er input i B4, og B5 lar deg skifte alle "
        "ratene med et påslag i prosentpoeng); prisene holdes flate, som i "
        "deck slide 5 der olje er 70 USD/fat fra 2035 og gass 5,7 USD/MMBtu "
        "fra 2040; statsandelen holdes på gjennomsnittet av de samme siste "
        "årene. Ratene er formler, så endres vinduet, endres hele utvidelsen. "
        "Halen er en regneteknisk forlengelse av en fallende bane, ikke en "
        "prognose — og konklusjonen mot 4 800 er robust for valget, se noten "
        "nederst.")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 88

    sr, sj = f"$B${E_VINDU}", f"$B${E_RJUST}"
    omr = lambda kol: _fu(f"${kol}$17:${kol}$42")   # 2025-2050
    rater = {E_ROLJE: ("B", "råolje"), E_RGASS: ("C", "naturgass"),
             E_RNGL: ("D", "NGL/kondensat"), E_RTOTH: ("F", "totalbane høy (SD)"),
             E_RTOTL: ("G", "totalbane lav (SD)"), E_RKOST: ("M", "kostnader")}
    par = [
        (E_VINDU, "Vindu for nedgangsrate (år)", 5,
         "Input. Ratene regnes over de siste så mange årene av basisbanen."),
        (E_RJUST, "Påslag på alle nedgangsrater (prosentpoeng)", 0.0,
         "Input for følsomhet. 0 = de observerte ratene. Et påslag på om lag "
         "3,4 prosentpoeng gir tilnærmet flat oljeproduksjon. Merk at et "
         "FELLES påslag ikke kan nulle ut rater som er ulike: ved +3,4 pp "
         "begynner kostnadene å vokse svakt."),
    ]
    for rad, (kol, navn) in rater.items():
        par.append((rad, f"Årlig endring, {navn}", None,
                    f"Avledet av Forutsetninger kolonne {kol}, pluss påslaget "
                    f"i B{E_RJUST}."))
    par += [
        (E_ANDEL, "Statsandel 2051-2060", None,
         "Gjennomsnitt av statsandelen over samme vindu. Andelen er et "
         "kalibreringsresidual og svinger 0,96-1,01, så et gjennomsnitt er "
         "mer robust enn 2050-verdien alene."),
        (E_PM, "PM-referanse, NNV 2026-2060 med 4 pst. (mrd.)", 4800,
         "Deck slide 5: «I PM var NNV i 2025 av kontantstrømmen 2026-2060 "
         "med 4 pst rente: 4800 mrd.» Faste priser med statsbudsjettets "
         "utgiftsdeflator."),
        (E_DEFL, "Deflatorjustering PM → modell", 1,
         "Modellen er i faste 2026-kroner. Er PM-tallet i 2025-kroner, må det "
         "ganges opp med ett års utgiftsdeflator (om lag 1,03) før "
         "sammenligning. Står på 1 = ingen justering, siden prisbasisen i "
         "PM-tallet ikke er verifisert. Merk at justeringen ØKER PM-tallet og "
         "dermed utvider differansen."),
        (E_NB26, "NB26s formuesberegning, statens del, 2026-2090, 3 pst.", 4721,
         "Input, uavhengig kontrollpunkt for halen."),
        (E_HALE, "Implisert verdi av halen 2051-2090, 3 pst.",
         f"=$B${E_NB26}-B41",
         "NB26s tall minus modellens NNV 3 pst. for 2026-2050. Brukes til å "
         "vurdere om ekstrapoleringen er for stram eller for romslig."),
    ]
    inputrader = {E_VINDU, E_RJUST, E_PM, E_DEFL, E_NB26}
    for rad, navn, verdi, note in par:
        ws.cell(rad, 1, navn)
        if rad in rater:
            kol = rater[rad][0]
            verdi = (f"=({_fu(f'${kol}$42')}/INDEX({omr(kol)},26-{sr}))"
                     f"^(1/{sr})-1+{sj}")
        elif rad == E_ANDEL:
            verdi = (f"=AVERAGEIF({_fu('$A$17:$A$42')},\">=\"&(2051-{sr}),"
                     f"{_fu('$P$17:$P$42')})")
        c = ws.cell(rad, 2, verdi)
        if rad in inputrader:
            c.font = BLA
        c.number_format = ("0.00 %" if rad in rater or rad == E_RJUST else
                           "0" if rad == E_VINDU else F2)
        ws.cell(rad, 3, note)

    # --- Årstabell 2051-2060, speiler Forutsetninger ---
    hdr = ["År", "Prod. råolje (mill. Sm3 o.e.)", "Prod. naturgass",
           "Prod. NGL/kond.", "Sum basis", "Total høy (SD, ekstrap.)",
           "Total lav (SD, ekstrap.)", "Volumfaktor høy", "Volumfaktor lav",
           "Oljepris (kr/Sm3 o.e.)", "Gasspris (kr/Sm3 o.e.)",
           "NGL-pris (kr/Sm3 o.e.)", "Kostnader (mill. kr)",
           "Basis SNCF (mill. kr)", "NKS modell (mill. kr)", "Statsandel"]
    for c, h in enumerate(hdr, 1):
        cc = ws.cell(EXT0 - 1, c, h)
        cc.font, cc.fill = HDR, HDRF
    vekst = {2: E_ROLJE, 3: E_RGASS, 4: E_RNGL, 6: E_RTOTH, 7: E_RTOTL,
             13: E_RKOST}
    for i in range(NYE):
        rad = EXT0 + i
        ws.cell(rad, 1, YEARS_EXT[i])
        for kol, prad in vekst.items():
            L = CL(kol)
            forrige = _fu(f"{L}42") if i == 0 else f"{L}{rad - 1}"
            ws.cell(rad, kol, f"={forrige}*(1+$B${prad})").number_format = F2
        ws.cell(rad, 5, f"=B{rad}+C{rad}+D{rad}").number_format = F2
        ws.cell(rad, 8, f"=F{rad}/E{rad}").number_format = F2
        ws.cell(rad, 9, f"=G{rad}/E{rad}").number_format = F2
        for kol in (10, 11, 12):        # priser flate på 2050-nivået
            ws.cell(rad, kol, f"={_fu(f'${CL(kol)}$42')}").number_format = F0
        ws.cell(rad, 16, f"=$B${E_ANDEL}").number_format = F2
        ws.cell(rad, 15, f"=B{rad}*J{rad}+C{rad}*K{rad}+D{rad}*L{rad}"
                         f"-M{rad}").number_format = F0
        ws.cell(rad, 14, f"=P{rad}*MAX(O{rad},"
                         f"({_fu(f'$B${R_GULV}')}-1)*1E+30)").number_format = F0

    # --- Samlet basisbane 2026-2060 i mrd., som NPV kan regne over ---
    ws.cell(E_BAS0 - 1, 18, "Samlet basisbane").font = FET
    ws.cell(E_BAS0 - 1, 19, "Mrd. 2026-kr").font = FET
    for i in range(NYA):
        rad, ref = E_BAS0 + i, _aarskilde(i)
        ws.cell(rad, 18, YEARS_ALL[i])
        ws.cell(rad, 19, f"={ref('N')}/1000").number_format = F0
    b50 = f"$S${E_BAS0}:$S${E_BAS0 + NY - 1}"
    b60 = f"$S${E_BAS0}:$S${E_BAS0 + NYA - 1}"

    # --- Fordelinger 2026-2060 ---
    ws.cell(31, 1, "NÅVERDI OG KUMULATIV 2026-2060 (mrd. 2026-kroner, "
                   "neddiskontert til 2025)").font = FET
    for c, h in enumerate(["", "P10", "P25", "P50", "P75", "P90", "Middel",
                           "Basisbane"], 1):
        if h:
            cc = ws.cell(32, c, h)
            cc.font, cc.fill = HDR, HDRF
    kv = (0.1, 0.25, 0.5, 0.75, 0.9)
    linjer = [
        ("NNV 4 pst., medianforankret", "NPV4_60", A_MED,
         f"=NPV({_fu('$B$10')},{b60})"),
        ("NNV 4 pst., forventningsforankret", "NPV4_60", A_FOR, None),
        ("NNV 3 pst., medianforankret", "NPV3_60", A_MED,
         f"=NPV({_fu('$B$8')},{b60})"),
        ("NNV 3 pst., forventningsforankret", "NPV3_60", A_FOR, None),
        ("Kumulativ udiskontert, medianforankret", "KUM60", A_MED,
         f"=SUM({b60})"),
        ("Kumulativ udiskontert, forventningsforankret", "KUM60", A_FOR, None),
    ]
    for k, (navn, nokkel, base, basisf) in enumerate(linjer):
        rad = 33 + k
        ws.cell(rad, 1, navn)
        col = CL(base + AGG.index(nokkel))
        rng = f"'{MOTOR}'!{col}{DATA0}:{col}{DATA0 + n - 1}"
        for j, q in enumerate(kv):
            ws.cell(rad, 2 + j, f"=PERCENTILE({rng},{q})").number_format = F0
        ws.cell(rad, 7, f"=AVERAGE({rng})").number_format = F0
        if basisf:
            ws.cell(rad, 8, basisf).number_format = F0

    # --- Broen fra hovedmodellen til PM-referansen ---
    ws.cell(40, 1, "BRO FRA HOVEDMODELLEN TIL PM-REFERANSEN (basisbanen)"
            ).font = FET
    bro = [
        ("1. Basis NNV 3 pst., 2026-2050 (hovedmodellen)",
         f"=NPV({_fu('$B$8')},{b50})", F0),
        ("2. Effekt av rente 3 → 4 pst., samme horisont",
         f"=NPV({_fu('$B$10')},{b50})-NPV({_fu('$B$8')},{b50})", F0),
        ("3. Effekt av horisont 2051-2060, ved 4 pst.",
         f"=NPV({_fu('$B$10')},{b60})-NPV({_fu('$B$10')},{b50})", F0),
        ("4. = Modellens basis NNV 2026-2060, 4 pst.", "=B41+B42+B43", F0),
        ("5. PM-referanse, deflatorjustert", f"=$B${E_PM}*$B${E_DEFL}", F0),
        ("6. Differanse, modell minus PM", "=B44-B45", F0),
        ("7. Differanse i pst. av PM", "=B46/B45", "0.0 %"),
    ]
    for k, (navn, f, fmt) in enumerate(bro):
        rad = 41 + k
        ws.cell(rad, 1, navn)
        ws.cell(rad, 2, f).number_format = fmt

    ws.cell(49, 1, "Kontroller").font = FET
    ws.cell(50, 1, "Bro rad 4 minus basiskolonnen over (skal være 0)")
    ws.cell(50, 2, "=B44-H33").number_format = F0
    ws.cell(51, 1, "Modellens hale 2051-2060, 3 pst.")
    ws.cell(51, 2, f"=NPV({_fu('$B$8')},{b60})-NPV({_fu('$B$8')},{b50})"
            ).number_format = F0
    ws.cell(52, 1, f"NB26s impliserte hale 2051-2090, 3 pst. (B{E_HALE})")
    ws.cell(52, 2, f"=$B${E_HALE}").number_format = F0

    ws.cell(54, 1,
            "Om sammenligningen: begge tall er nåverdier datert 2025 — Excels "
            "NPV() diskonterer første beløp ett år, og PM-tallet er oppgitt "
            "som «NNV i 2025». Broen viser at horisont og rente nesten "
            "opphever hverandre: å gå fra 3 til 4 pst. koster mer enn de tio "
            "ekstra årene tilfører. Differansen mot 4 800 ligger derfor i "
            "kontantstrømmens NIVÅ, ikke i horisonten eller renten, og den var "
            "der allerede på 2026-2050 med 3 pst."
            ).font = GRA
    ws.cell(55, 1,
            "Robusthet: et påslag i B5 fra 0 til +5 prosentpoeng på alle "
            "nedgangsrater flytter NNV 4 pst. bare fra 3 663 til 3 716 mrd. "
            "Halen er liten og ligger langt ute i diskonteringen, så "
            "konklusjonen om et gap på om lag 1 100 mrd. er ikke et resultat "
            "av ekstrapoleringsvalget."
            ).font = GRA
    ws.cell(56, 1,
            "Halekontrollen i rad 51-52 peker samme vei: NB26s egen "
            "formuesberegning impliserer at halen etter 2050 er verdt klart "
            "mer enn modellens forlengelse gir. Modellens hale er altså om "
            "noe konservativ, så heller ikke den forklarer differansen mot "
            "PM-tallet. Gjenstående kandidater er årgang (PM mot NB26), "
            "prisbasis, og hva PM regner med som ikke ligger i NB26-grunnlaget."
            ).font = GRA
    ws.cell(57, 1,
            "Basisbanen er ikke det samme som medianen i viften. "
            "Sammenligningen mot 4 800 gjøres på basisbanen; fordelingen i "
            "rad 33-38 viser hva usikkerheten gjør med det samme tallet."
            ).font = GRA

    ws.column_dimensions["A"].width = 46
    for col in range(2, 20):
        ws.column_dimensions[CL(col)].width = 13
    ws.column_dimensions["C"].width = 58
    return ws


def _vifte(wb, n=N):
    """Synlig ark: persentiler for begge forankringer, basis og NZE-scenario."""
    if VIFTE in wb.sheetnames:
        del wb[VIFTE]
    ws = wb.create_sheet(VIFTE, 4)

    ws["A1"] = ("REFORMULERT USIKKERHETSMODELL — SNCF, mrd. 2026-kroner. "
                "Persentilene er scenariene.")
    ws["A1"].font = FET
    ws["A2"] = (
        f"Én Monte Carlo på basisproduksjon, {n} simuleringer med faste trekk "
        "(frø i Forutsetninger). Pris og volum trekkes som PERSISTENTE REGIMER "
        "— ett trekk per simulering som gjelder hele perioden — fordi høy- og "
        "lavbaner historisk har vart over flere år. Pris: korrelerte lognormale "
        "faktorer på NB26-banen, sigma avledet av de historiske "
        "persentilforholdene i Forutsetninger B4:B7. Volum: triangulær vekt som "
        "interpolerer Sokkeldirektoratets lav/basis/høy. Feltnetto gulves ved 0 "
        "(balanseprisgulv). BEGGE forankringer vises på samme trekk: "
        "medianforankring har P50 = NB26 og et middel over NB26; "
        "forventningsforankring har middel = NB26 og en median under. "
        "Middelet skal ikke rapporteres som forventet innbetaling uten at "
        "avviket mot NB26 opplyses.")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 78

    grupper = [(2, "Medianforankring (hovedspor): P50 = NB26"),
               (8, "Forventningsforankring (følsomhet): middel = NB26"),
               (14, "Referansebaner")]
    for col, tekst in grupper:
        c = ws.cell(3, col, tekst)
        c.font, c.fill = HDR, HDRF
    kolonner = ["År"] + ["P10", "P25", "P50", "P75", "P90", "Middel"] * 2 \
        + ["NB26-basis", "IEA NZE-sidescenario"]
    for c, h in enumerate(kolonner, 1):
        cc = ws.cell(4, c, h)
        cc.font, cc.fill = HDR, HDRF

    kv = [("P10", 0.1), ("P25", 0.25), ("P50", 0.5), ("P75", 0.75), ("P90", 0.9)]
    for i in range(NY):
        row, fr = 5 + i, FU0 + i
        ws.cell(row, 1, YEARS[i])
        for j, (base, kol0) in enumerate(((C_MED, 2), (C_FOR, 8))):
            rng = f"'{MOTOR}'!{CL(base + i)}{DATA0}:{CL(base + i)}{DATA0 + n - 1}"
            for k, (_, q) in enumerate(kv):
                ws.cell(row, kol0 + k, f"=PERCENTILE({rng},{q})").number_format = F0
            ws.cell(row, kol0 + 5, f"=AVERAGE({rng})").number_format = F0
        # NB26-basis som formel, ikke hardkodet tall.
        ws.cell(row, 14, f"={_fu(f'N{fr}')}/1000").number_format = F0
        ws.cell(row, 15, _nze_formel(fr)).number_format = F0

    ws.cell(31, 1, "Oppsummering (mrd. 2026-kroner)").font = FET
    for col, tekst in ((2, "Medianforankring"), (8, "Forventningsforankring")):
        c = ws.cell(32, col, tekst)
        c.font, c.fill = HDR, HDRF
    for c, h in enumerate(["", "P10", "P25", "P50", "P75", "P90", "Middel"] * 2, 1):
        if h:
            cc = ws.cell(33, c, h)
            cc.font, cc.fill = HDR, HDRF
    for k, (navn, off) in enumerate((("Kumulativ SNCF 2026-2050", 0),
                                     ("NPV, 3 pst.", 1),
                                     ("NPV, 2 pst.", 2),
                                     ("NPV, 4 pst.", 3))):
        row = 34 + k
        ws.cell(row, 1, navn)
        for base, kol0 in ((A_MED, 2), (A_FOR, 8)):
            col = CL(base + off)
            rng = f"'{MOTOR}'!{col}{DATA0}:{col}{DATA0 + n - 1}"
            for j, (_, q) in enumerate(kv):
                ws.cell(row, kol0 + j, f"=PERCENTILE({rng},{q})").number_format = F0
            ws.cell(row, kol0 + 5, f"=AVERAGE({rng})").number_format = F0

    ws.cell(39, 1, "Kontroll mot basis").font = FET
    kontroller = [
        ("NB26-basis, kumulativ 2026-2050", f"=SUM(N5:N{4 + NY})"),
        ("NB26-basis, NPV 3 pst.", f"=NPV({_fu('$B$8')},N5:N{4 + NY})"),
        ("Median P50 kumulativ / basis - 1", "=D34/B40-1"),
        ("Median middel kumulativ / basis - 1", "=G34/B40-1"),
        ("Forventning middel kumulativ / basis - 1", "=M34/B40-1"),
        ("Sum av årsmedianer (medianforankring)", f"=SUM(D5:D{4 + NY})"),
    ]
    for k, (navn, f) in enumerate(kontroller):
        row = 40 + k
        ws.cell(row, 1, navn)
        c = ws.cell(row, 2, f)
        c.number_format = "0.0 %" if "- 1" in navn else F0

    ws.cell(47, 1, "Parametre og forankring").font = FET
    par = [
        ("sigma_olje (avledet, historisk)", f"={_fu(f'B{R_SIGO}')}", F2),
        ("sigma_gass (avledet, historisk)", f"={_fu(f'B{R_SIGG}')}", F2),
        ("korrelasjon olje/gass", f"={_fu(f'B{R_RHO}')}", F2),
        ("Balansepris-gulv (1 = på)", f"={_fu(f'B{R_GULV}')}", "0"),
        ("Antall simuleringer", f"={_fu(f'B{R_N}')}", "0"),
        ("Frø (seed)", f"={_fu(f'B{R_SEED}')}", "0"),
        ("Kalibrering", "historisk (Forutsetninger B4:B7), vei (c)", None),
        ("NZE olje faller på persentil", f"={_fu(f'B{R_PNZEO}')}", "0.0 %"),
        ("NZE gass faller på persentil", f"={_fu(f'B{R_PNZEG}')}", "0.0 %"),
        ("STEPS olje faller på persentil", f"={_fu(f'B{R_PSTPO}')}", "0.0 %"),
        ("STEPS gass faller på persentil", f"={_fu(f'B{R_PSTPG}')}", "0.0 %"),
    ]
    for k, (navn, f, fmt) in enumerate(par):
        row = 48 + k
        ws.cell(row, 1, navn)
        c = ws.cell(row, 2, f)
        if fmt:
            c.number_format = fmt

    ws.cell(60, 1,
            "IEA-scenariene er IKKE brukt til å kalibrere viften. Grunnen er at "
            "de skiller seg ved politikk og etterspørsel, ikke ved tilbudssjokk, "
            "og derfor nesten bare spenner nedsiden: STEPS ligger tett på APS "
            "mens NZE ligger langt under. Å tvinge P90 = STEPS og P10 = NZE "
            "krever sigma som er over ti ganger ulike opp og ned, og gir en "
            "vifte uten mening. NZE vises derfor som et navngitt sidescenario, "
            "og persentilavlesningen over sier hvor uenige kildene er."
            ).font = GRA
    ws.cell(62, 1,
            f"Horisonten her er 2026-2050, som er så langt "
            f"Sokkeldirektoratets mulighetsbilder rekker. Utvidelsen til 2060 "
            f"og sammenligningen mot PM-referansen på 4 800 mrd. ligger i "
            f"arket «{UTV}», der 2051-2060 er ekstrapolert."
            ).font = GRA
    ws.cell(61, 1,
            "NZE-kolonnen gir #N/A til prisene er lagt inn i Forutsetninger "
            f"(B{R_NZEO}/B{R_NZEG} for flatt nivå, eller kolonne Q/R for en "
            "årsbane). Tallene skal hentes fra IEA WEO Annex A."
            ).font = GRA

    ws.column_dimensions["A"].width = 38
    for col in range(2, 16):
        ws.column_dimensions[CL(col)].width = 12
    ws.freeze_panes = "B5"
    return ws


def _nze_formel(fr):
    """NZE-sidescenario for ett år: basisvolum, NZE-priser, samme gulv.

    NGL følger oljeprisen, som i resten av modellen. Prisene skaleres som
    forhold mot basisbanen, så enhetene arves fra Forutsetninger.
    """
    o_usd = (f"IF(ISNUMBER({_fu(f'Q{fr}')}),{_fu(f'Q{fr}')},"
             f"{_fu(f'$B${R_NZEO}')})")
    g_usd = (f"IF(ISNUMBER({_fu(f'R{fr}')}),{_fu(f'R{fr}')},"
             f"{_fu(f'$B${R_NZEG}')})")
    o_kr = f"{o_usd}*{_fu(f'$B${R_BBL}')}*{_fu(f'$B${R_FX}')}"
    g_kr = f"{g_usd}*{_fu(f'$B${R_KRG}')}"
    ro = f"({o_kr})/{_fu(f'J{fr}')}"
    rg = f"({g_kr})/{_fu(f'K{fr}')}"
    netto = (f"({_fu(f'B{fr}')}*{_fu(f'J{fr}')}+{_fu(f'D{fr}')}*{_fu(f'L{fr}')})"
             f"*{ro}+{_fu(f'C{fr}')}*{_fu(f'K{fr}')}*{rg}-{_fu(f'M{fr}')}")
    kropp = (f"{_fu(f'P{fr}')}*IF({_fu(f'$B${R_GULV}')}=1,"
             f"MAX({netto},0),{netto})/1000")
    return f"=IF({_fu(f'$B${R_NZEOK}')}=0,NA(),{kropp})"


MARKOR = ("REFORMULERT USIKKERHETSMODELL (arket «Reformulert vifte» "
          "og skjult «MC-motor-R»)")


def _dokumentasjon(wb):
    """Skriver dokumentasjonsblokken. Idempotent: finner markøren fra forrige
    kjøring og skriver over, i stedet for å legge til en ny kopi."""
    dok = wb["Dokumentasjon"]
    rad = None
    for r in range(1, dok.max_row + 1):
        if dok.cell(r, 1).value == MARKOR:
            rad = r
            break
    if rad is None:
        rad = dok.max_row + 2
    else:
        for r in range(rad, dok.max_row + 1):
            dok.cell(r, 1, None)
    dok.cell(rad, 1, MARKOR).font = FET
    tekster = [
        "Persentilene er scenariene: P90 er den høye prognosen, P50 "
        "sentralanslaget og P10 den lave. Den erstatter 3x3-en med vilkårlige "
        "multiplikatorer, som ble forlatt fordi høy/lav-nivåene ikke hadde "
        "kilde og fordi kombinasjonen høyt volum og lav pris ga negative tall.",
        "Usikkerheten trekkes som PERSISTENTE REGIMER, ett trekk per "
        "simulering for hele perioden, ikke som år-til-år-støy. Det speiler at "
        "høy- og lavprisperioder historisk har vart i flere år, og gir en "
        "vifte som kan leses som scenarier.",
        "Tilbudsrespons: feltnetto gulves ved 0 (bryter i Forutsetninger). "
        "Uten gulvet er om lag 11 pst. av årsobservasjonene negative, som "
        "overdriver nedsiden — produsentene stenger ned i stedet for å "
        "produsere med tap. Gulvet er forankret i balanseprisene, deck slide "
        "14 (om lag 20-45 USD/fat før skatt).",
        "BEGGE forankringer vises, på samme trekk. Medianforankring "
        "(hovedspor) setter median[pris] = NB26, slik at P50-banen ER "
        "sentralanslaget; til gjengjeld ligger middelet over NB26. "
        "Forventningsforankring setter E[pris] = NB26 via Jensen-korreksjonen "
        "EXP(-sigma^2/2); da er middelet lik NB26, men medianen faller under. "
        "Begge kan ikke oppfylles samtidig for en skjev fordeling. Middelet "
        "skal ikke presenteres som forventet innbetaling til fondet uten at "
        "avviket mot NB26 opplyses.",
        "Kalibrering: sigma avledes av de re-sentrerte historiske "
        "persentilforholdene (Forutsetninger B4:B7) som "
        "LN(forhold)/NORMSINV(0,9). Det er eksakt for en medianforankret "
        "lognormal, og gjør viften konsistent med prisskiftene i «Statisk "
        "modell» i P10/P90. Ingenting er hardkodet: endres B4:B7, endres sigma.",
        "IEA WEO er IKKE brukt til å kalibrere viftens ytterkanter. "
        "Scenariene skiller seg ved politikk og etterspørsel, ikke ved "
        "tilbudssjokk, og spenner derfor nesten bare nedsiden — IEA har ingen "
        "høyprisverden. NZE er i stedet lagt inn som et navngitt sidescenario "
        "utenfor viften, slik at markedsrisiko (viften, kalibrert på "
        "historikk) og politikkrisiko (NZE-linjen) holdes fra hverandre.",
        "UTVIDELSE TIL 2060 (arket «Utvidelse 2060»): horisonten i "
        "hovedmodellen er 2026-2050, som er så langt Sokkeldirektoratets "
        "mulighetsbilder rekker. Utvidelsen forlenger volumer, totalbaner og "
        "kostnader med de geometriske ratene fra de siste årene av "
        "basisbanen, holder prisene flate og holder statsandelen på snittet "
        "av samme vindu. Prisene er flate allerede fra 2041 i basisbanen, så "
        "det er volumfallet som driver halen. Årene etter 2050 har ingen "
        "kilde og er markert som ekstrapolert i figurene.",
        "Sammenligningen mot PM-referansen: deck slide 5 oppgir at "
        "nåverdien i 2025 av kontantstrømmen 2026-2060 med 4 pst. rente var "
        "4 800 mrd. i Perspektivmeldingen. Modellens basisbane gir 3 663 mrd. "
        "på samme horisont og rente. Broen i arket viser at horisont og rente "
        "nesten opphever hverandre — overgangen fra 3 til 4 pst. koster mer "
        "enn de ti ekstra årene tilfører — så differansen på om lag 1 100 "
        "mrd. ligger i kontantstrømmens nivå, ikke i regnemåten. Den var "
        "der allerede på 2026-2050 med 3 pst. Tallene skal derfor ikke "
        "presenteres som samme størrelse.",
        "Halekontrollen: NB26s egen formuesberegning for statens del "
        "2026-2090 med 3 pst. er 4 721 mrd. Trekker man fra modellens "
        "3 753 mrd. for 2026-2050, impliserer NB26 at halen etter 2050 er "
        "verdt om lag 968 mrd. Modellens ekstrapolerte hale er klart "
        "lavere, og selv en flat produksjon på 2050-nivået helt til 2090 "
        "kommer ikke opp til NB26s tall. Ekstrapoleringen er altså om noe "
        "konservativ, og forklarer ikke differansen mot PM-tallet.",
        "Forbehold: den historiske kalibreringen bygger på 1997-2024, en "
        "periode uten gjennomført energiomstilling. Persentilavlesningen i "
        "arket viser hvor et NZE-forløp havner i viften — ligger det svært "
        "langt ute i halen, betyr det at viften ikke rommer "
        "omstillingsrisikoen, og NZE-linjen må leses som et selvstendig "
        "scenario, ikke som en usannsynlig hale.",
    ]
    for k, t in enumerate(tekster, 1):
        dok.cell(rad + k, 1, t)


def build(fil=FIL, n=N, seed=SEED):
    wb = load_workbook(fil)
    w, z1, z2 = trekk(n, seed)
    _parametre(wb)
    _utvidelse(wb, n)
    _motor(wb, w, z1, z2, n)
    _vifte(wb, n)
    _dokumentasjon(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.save(fil)
    return fil


if __name__ == "__main__":
    print("Skrev", build())
