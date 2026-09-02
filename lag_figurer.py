# -*- coding: utf-8 -*-
"""
Regenererer de tre SVG-figurene fra den forventningsforankrede modellen.

Leser de faste trekkene (w, z1, z2, frø 2026) og forutsetningene direkte fra
Kontantstromsmodell_petroleum.xlsx og beregner den forventningsforankrede
SNCF-en (samme modell som MC-motoren og mc_simulering.py), slik at figurene er
nøyaktig konsistente med arbeidsboken.

Figurer (FINs designprofil, hvit bakgrunn, ingen tittel/kilde i figuren):
  viftefigur_sncf.svg   - tetthetsvifte for årlig SNCF (P5-P95), median og
                          NB26-basis.
  akkumulert_sncf.svg   - akkumulert SNCF med persentilbånd og sluttverdier.
  fordelinger_sncf.svg  - histogrammer for SNCF 2035 og kumulativ, med modus,
                          median, middelverdi og basisbane.

Kjøring:  python3 lag_figurer.py
Krever:   numpy, openpyxl, matplotlib
"""
import sys
import numpy as np
from openpyxl import load_workbook

# FINs designprofil (matplotlib-stil)
sys.path.insert(0, "/root/.claude/skills/synced/"
                "5030c864-d034-405a-808f-f52ea4fe255a_9e3f4a54-160c-4203-91ed-822663bb2dc7/"
                "fin-designprofil/scripts")
from fin_chart_style import (apply_matplotlib_style, FIN_DARKBLUE, FIN_BLUE,
                             FIN_LIGHTBLUE, FIN_RED)
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

apply_matplotlib_style()
FIL = "Kontantstromsmodell_petroleum.xlsx"
YEARS = np.arange(2026, 2051)


def nf(x, dec=0):
    """Norsk tallformat: mellomrom som tusenskille, komma som desimaltegn."""
    s = f"{x:,.{dec}f}".replace(",", " ").replace(".", ",")
    return s


def simuler():
    """Forventningsforankret SNCF over de 2 000 faste trekkene i arket."""
    wb = load_workbook(FIL, data_only=True)
    fu = wb["Forutsetninger"]
    sO, sG, rho = fu["B11"].value, fu["B12"].value, fu["B13"].value
    kO, kG = fu["B14"].value, fu["B15"].value
    col = lambda L: np.array([fu[f"{L}{18+i}"].value for i in range(25)])
    volO, volG, volN = col("B"), col("C"), col("D")
    totH, totL = col("F"), col("G")
    pO, pG, pN, cost, snks = col("J"), col("K"), col("L"), col("M"), col("N")
    tot = volO + volG + volN
    fh, fl = totH / tot, totL / tot
    nks = volO * pO + volG * pG + volN * pN - cost
    andel = snks / nks
    snks_b = andel * nks / 1000.0                     # NB26-basis per år (mrd.)

    mc = wb["MC-motor"]
    n = 2000
    w = np.array([mc.cell(10 + r, 2).value for r in range(n)])
    z1 = np.array([[mc.cell(10 + r, 3 + i).value for i in range(25)] for r in range(n)])
    z2 = np.array([[mc.cell(10 + r, 28 + i).value for i in range(25)] for r in range(n)])

    phiO, phiG = 1 - kO, 1 - kG
    volfac = np.where(w[:, None] >= 0, 1 + w[:, None] * (fh - 1),
                      1 - (-w[:, None]) * (1 - fl)) / (1 + (fh + fl - 2) / 6)
    epsO, epsG = z1, rho * z1 + np.sqrt(1 - rho ** 2) * z2
    lMo = np.zeros((n, 25)); lMg = np.zeros((n, 25))
    vO = np.zeros(25); vG = np.zeros(25); a = b = 0.0
    for t in range(25):
        lMo[:, t] = (lMo[:, t - 1] * phiO if t else 0) + sO * epsO[:, t]
        lMg[:, t] = (lMg[:, t - 1] * phiG if t else 0) + sG * epsG[:, t]
        a = phiO ** 2 * a + sO ** 2; b = phiG ** 2 * b + sG ** 2
        vO[t] = a; vG[t] = b
    Mo, Mg = np.exp(lMo - 0.5 * vO), np.exp(lMg - 0.5 * vG)
    rev = volfac * (volO * pO * Mo + volG * pG * Mg + volN * pN * Mo)
    sncf = andel * (rev - volfac * cost) / 1000.0     # (n, 25) mrd. 2026-kroner
    return sncf, snks_b


def _fin_akser(ax):
    """Y-akse på begge sider, tickmarks innover, ingen x-tickmarks."""
    ax.tick_params(axis="y", which="both", direction="in", right=True, left=True,
                   labelright=False, length=3)
    ax.tick_params(axis="x", length=0)
    for s in ax.spines.values():
        s.set_linewidth(0.5); s.set_color("#000000")
    ax.spines["top"].set_visible(False)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: nf(v)))


def figur_vifte(sncf, snks_b):
    fig, ax = plt.subplots(figsize=(8.9, 4.72))
    med = np.median(sncf, axis=0)
    # Gradert tetthetsvifte, kuttet ved P5-P95
    lows = np.arange(5, 50, 5)
    for lo in lows:
        hi = 100 - lo
        band_lo = np.percentile(sncf, lo, axis=0)
        band_hi = np.percentile(sncf, hi, axis=0)
        alpha = 0.10 + 0.55 * (lo / 45.0)             # tettere mot midten
        ax.fill_between(YEARS, band_lo, band_hi, color=FIN_BLUE,
                        alpha=alpha, linewidth=0, zorder=1)
    ax.plot(YEARS, snks_b, color=FIN_RED, ls="--", lw=1.5,
            label="Basisbane (NB26)", zorder=4)
    ax.plot(YEARS, med, color=FIN_DARKBLUE, lw=3.0, zorder=5)   # mørk kontur
    ax.plot(YEARS, med, color="white", lw=1.4, zorder=6, label="Median")
    ax.axhline(0, color="#000000", lw=0.5, zorder=2)
    ax.set_xlim(2026, 2050); ax.margins(x=0)
    ax.set_xticks([2030, 2035, 2040, 2045, 2050])
    _fin_akser(ax)
    ax.annotate("Mrd. 2026-kroner", xy=(0, 1.02), xycoords="axes fraction",
                fontsize=9, color="#000000")
    ax.legend(loc="upper right", fontsize=9)
    fig.tight_layout()
    fig.savefig("viftefigur_sncf.svg"); plt.close(fig)


def figur_akkumulert(sncf, snks_b):
    fig, ax = plt.subplots(figsize=(8.9, 4.72))
    cum = np.cumsum(sncf, axis=1)
    P = {q: np.percentile(cum, q, axis=0) for q in [10, 25, 50, 75, 90]}
    ax.fill_between(YEARS, P[10], P[90], color=FIN_LIGHTBLUE, alpha=0.35,
                    linewidth=0, label="10-90-persentil")
    ax.fill_between(YEARS, P[25], P[75], color=FIN_BLUE, alpha=0.45,
                    linewidth=0, label="25-75-persentil")
    ax.plot(YEARS, P[50], color=FIN_DARKBLUE, lw=1.8, label="Median")
    for q in [10, 25, 50, 75, 90]:
        ax.annotate(f"P{q}: {nf(P[q][-1])}", xy=(2050, P[q][-1]),
                    xytext=(6, 0), textcoords="offset points", va="center",
                    fontsize=8, color="#000000")
    ax.set_xlim(2026, 2050); ax.margins(x=0)
    ax.set_xticks([2030, 2035, 2040, 2045, 2050])
    ax.set_xlim(2026, 2053)                            # plass til sluttverdier
    _fin_akser(ax)
    ax.annotate("Mrd. 2026-kroner, akkumulert fra 2026", xy=(0, 1.02),
                xycoords="axes fraction", fontsize=9, color="#000000")
    ax.legend(loc="upper left", fontsize=9)
    fig.tight_layout()
    fig.savefig("akkumulert_sncf.svg"); plt.close(fig)


def _hist_panel(ax, data, basis, tittel):
    from matplotlib.lines import Line2D
    counts, edges, _ = ax.hist(data, bins=40, color=FIN_LIGHTBLUE,
                               edgecolor="white", linewidth=0.3)
    mid = 0.5 * (edges[:-1] + edges[1:])
    modus = mid[np.argmax(counts)]
    med = np.median(data); mean = data.mean()
    # Markørene ligger tett; bruk tegnforklaring med verdier i stedet for
    # overlappende tekst inne i figuren.
    marks = [("Modus (flest utfall)", modus, FIN_DARKBLUE, ":"),
             ("Median", med, FIN_DARKBLUE, "-"),
             ("Middelverdi", mean, FIN_BLUE, "-"),
             ("Basisbane (NB26)", basis, FIN_RED, "--")]
    handles = []
    for lbl, x, color, ls in marks:
        ax.axvline(x, color=color, ls=ls, lw=1.5)
        handles.append(Line2D([0], [0], color=color, ls=ls, lw=1.5,
                              label=f"{lbl}: {nf(x)}"))
    ax.legend(handles=handles, loc="upper right", fontsize=8,
              handlelength=1.6, borderaxespad=0.4)
    ax.set_xlabel(tittel, fontsize=9, color="#000000")
    ax.tick_params(axis="y", direction="in", left=True, right=True, length=3)
    ax.tick_params(axis="x", direction="in", length=3)
    for s in ax.spines.values():
        s.set_linewidth(0.5); s.set_color("#000000")
    ax.spines["top"].set_visible(False)
    ax.set_yticks([])
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: nf(v)))


def figur_fordelinger(sncf, snks_b):
    fig, (axL, axR) = plt.subplots(1, 2, figsize=(10.38, 4.10))
    i2035 = list(YEARS).index(2035)
    _hist_panel(axL, sncf[:, i2035], snks_b[i2035],
                "SNCF i 2035 (mrd. 2026-kroner)")
    _hist_panel(axR, np.cumsum(sncf, axis=1)[:, -1], snks_b.sum(),
                "Akkumulert SNCF 2026-2050 (mrd. 2026-kroner)")
    fig.tight_layout()
    fig.savefig("fordelinger_sncf.svg"); plt.close(fig)


if __name__ == "__main__":
    sncf, snks_b = simuler()
    cum = np.cumsum(sncf, axis=1)[:, -1]
    figur_vifte(sncf, snks_b)
    figur_akkumulert(sncf, snks_b)
    figur_fordelinger(sncf, snks_b)
    print("Skrev viftefigur_sncf.svg, akkumulert_sncf.svg, fordelinger_sncf.svg")
    print(f"Kontroll: akkumulert P10/P50/P90 = "
          f"{nf(np.percentile(cum,10))}/{nf(np.percentile(cum,50))}/"
          f"{nf(np.percentile(cum,90))}  middel {nf(cum.mean())} (basis {nf(snks_b.sum())})")
