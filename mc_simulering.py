# -*- coding: utf-8 -*-
"""
Monte Carlo-simulering av statens netto kontantstrøm (SNCF) fra
petroleumsvirksomheten, 2026-2050. Faste 2026-kroner.

Leser forutsetningene direkte fra arket "Forutsetninger" i
Kontantstromsmodell_petroleum.xlsx, slik at endringer der slår
gjennom i simuleringen.

Modellvalg (jf. Dokumentasjon-arket):
- Volum: ett trekk per simulering; triangulær vekt w ~ Tri(-1, 0, 1)
  interpolerer mellom lav- (w=-1), basis- (w=0) og høybanen (w=+1).
- Pris: gjennomsnittsreverterende prosess (Ornstein-Uhlenbeck på
  logpris) FORVENTNINGSFORANKRET i NB26-banen (E[pris]=NB26 via
  Jensen-korreksjon), fordi NB26-banen er en forventningsbane, ikke en
  median. Kappa estimert med AR(1) på realpriser 1997-2024.
  OVERSTYR_KAPPA = 0.0 gir ren GBM.
- Volum: forventningsforankret, E[volfaktor]=1.
- NGL følger oljesjokket. Pris-volum-korrelasjon er null.

Kjøring:  python mc_simulering.py
Krever:   numpy, openpyxl
"""
import numpy as np
from openpyxl import load_workbook

# ----------------------- Parametre -----------------------
FIL = "Kontantstromsmodell_petroleum.xlsx"
N = 10_000          # antall simuleringer
SEED = 2026         # frø for reproduserbarhet
# Reverteringshastigheter leses fra arket (B14/B15). Sett OVERSTYR_KAPPA
# til et tall (f.eks. 0.0 for ren GBM) for å overstyre begge.
OVERSTYR_KAPPA = None
RENTER = [0.02, 0.03, 0.04]
PCT = [10, 25, 50, 75, 90]

# ----------------------- Innlesing -----------------------
wb = load_workbook(FIL, data_only=True)
ws = wb["Forutsetninger"]
sO, sG, rho = ws["B11"].value, ws["B12"].value, ws["B13"].value
kapO, kapG = ws["B14"].value, ws["B15"].value
if OVERSTYR_KAPPA is not None:
    kapO = kapG = OVERSTYR_KAPPA
phiO, phiG = 1 - kapO, 1 - kapG

R0, NYRS = 17, 26   # tabellstart (2025) og antall år
rows = list(ws.iter_rows(min_row=R0, max_row=R0 + NYRS - 1, values_only=True))
years = np.array([r[0] for r in rows])
volO = np.array([r[1] for r in rows]); volG = np.array([r[2] for r in rows])
volN = np.array([r[3] for r in rows])
totB = volO + volG + volN
totH = np.array([r[5] for r in rows]); totL = np.array([r[6] for r in rows])
pO = np.array([r[9] for r in rows]); pG = np.array([r[10] for r in rows])
pN = np.array([r[11] for r in rows])
cost = np.array([r[12] for r in rows]); snks = np.array([r[13] for r in rows])
nks = volO * pO + volG * pG + volN * pN - cost
andel = snks / nks

m = years >= 2026   # modellår (2025 er anker)
years, volO, volG, volN = years[m], volO[m], volG[m], volN[m]
pO, pG, pN, cost, andel = pO[m], pG[m], pN[m], cost[m], andel[m]
fh, fl = (totH / totB)[m], (totL / totB)[m]
T = len(years)

# ----------------------- Simulering -----------------------
rng = np.random.default_rng(SEED)

w = rng.triangular(-1, 0, 1, size=N)
volfac = np.where(w[:, None] >= 0,
                  1 + w[:, None] * (fh - 1),
                  1 - (-w[:, None]) * (1 - fl))
# Forventningsforankring, volum: del på E[volfaktor] = 1 + (fh + fl - 2)/6
# (lukket form for den symmetriske triangulære vekten), slik at E[volfac]=1.
volfac = volfac / (1 + (fh + fl - 2) / 6.0)

L = np.linalg.cholesky([[1, rho], [rho, 1]])
eps = rng.standard_normal((N, T, 2)) @ L.T

logMo = np.zeros((N, T)); logMg = np.zeros((N, T))
varO = np.zeros(T); varG = np.zeros(T); vo = vg = 0.0
for t in range(T):
    prev_o = logMo[:, t - 1] if t else 0.0
    prev_g = logMg[:, t - 1] if t else 0.0
    logMo[:, t] = prev_o * phiO + sO * eps[:, t, 0]
    logMg[:, t] = prev_g * phiG + sG * eps[:, t, 1]
    vo = phiO ** 2 * vo + sO ** 2      # Var(logMo_t) = phi^2*Var_{t-1} + sigma^2
    vg = phiG ** 2 * vg + sG ** 2
    varO[t] = vo; varG[t] = vg
# Forventningsforankring, pris (Jensen): E[pris]=NB26 => E[Mo]=E[Mg]=1.
# Korreksjonen ligger KUN på eksponentieringen, ikke i OU-rekursjonen over
# (ellers reverterer prosessen mot et senket nivå og feilen komponerer).
Mo, Mg = np.exp(logMo - 0.5 * varO), np.exp(logMg - 0.5 * varG)

rev = volfac * (volO * pO * Mo + volG * pG * Mg + volN * pN * Mo)
sncf = andel * (rev - volfac * cost) / 1000.0   # mrd. 2026-kroner

# ----------------------- Resultater -----------------------
disc = np.arange(1, T + 1)
print(f"Simuleringer: {N}, kappa olje/gass = {kapO:.3f}/{kapG:.3f}, frø = {SEED}")
print(f"{'År':<6}" + "".join(f"P{q:<8}" for q in PCT))
for i, y in enumerate(years):
    print(f"{y:<6}" + "".join(f"{np.percentile(sncf[:, i], q):<9.1f}" for q in PCT))
cum = sncf.sum(axis=1)
print("\nKumulativ SNCF 2026-2050 (mrd.):",
      {f"P{q}": round(float(np.percentile(cum, q))) for q in PCT})
for r in RENTER:
    npv = (sncf / (1 + r) ** disc).sum(axis=1)
    print(f"NPV {r:.0%}:",
          {f"P{q}": round(float(np.percentile(npv, q))) for q in PCT},
          f"middel: {npv.mean():.0f}")
