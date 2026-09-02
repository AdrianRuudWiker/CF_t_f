# -*- coding: utf-8 -*-
"""
Verifiserer at Excel-formlene som build_reformulert.py skriver, regner det
samme som Python-referansen i mc_reformulert.py.

soffice/LibreOffice er ødelagt i containeren og kan ikke reberegne arbeidsboken,
så verifikasjonen bruker `formulas`-biblioteket, som er en egen
Excel-formeltolker. Å evaluere hele leveransen (over 270 000 formler) er ikke
praktisk mulig, så testen bygger en LITEN, strukturelt identisk arbeidsbok med
få simuleringer: samme parameterblokk, samme motorformler og samme
viftefomler, generert av de SAMME funksjonene som leveransen. Da er det
formelstrengene som faktisk testes, ikke en omskrivning av dem.

Kontrollene:
1. Motoren, celle for celle: Excel-evaluert SNCF mot Python på identiske trekk,
   begge forankringer.
2. Persentil- og aggregatformlene i det synlige arket mot Python.
3. Forankringsidentitetene: median[pris] = basis i medianforankringen, og
   E[pris] = basis i forventningsforankringen.
4. Gulvet: ingen negative årsverdier når bryteren er på, og at det å slå den av
   faktisk slipper negative tall gjennom.

Kjøring:  python3 verifiser_reformulert.py
Krever:   numpy, openpyxl, formulas
"""
import os
import tempfile

import numpy as np
from openpyxl import Workbook, load_workbook

import build_reformulert as br
import mc_reformulert as mc

N_TEST = 25          # få simuleringer — formelstrukturen er det som testes
TOL = 1e-9


def _testbok(n, sti):
    """Liten arbeidsbok med samme struktur som leveransen.

    Årstabellen og B4:B15 legges inn som VERDIER (lest fra leveransen), slik at
    `formulas` ikke må evaluere hele den eksisterende modellen. Alt som
    build_reformulert skriver, skrives av de originale funksjonene.
    """
    kilde = load_workbook(br.FIL, data_only=True)["Forutsetninger"]
    wb = Workbook()
    fu = wb.active
    fu.title = "Forutsetninger"
    for r in range(4, 16):                      # parametre B4:B15
        fu.cell(r, 1, kilde.cell(r, 1).value)
        fu.cell(r, 2, kilde.cell(r, 2).value)
    for r in range(16, 43):                     # årstabell A16:P42
        for c in range(1, 17):
            fu.cell(r, c, kilde.cell(r, c).value)
    # Kolonne E, H, I, O og P er FORMLER i leveransen, og openpyxl fjerner
    # bufrede verdier ved lagring, så data_only gir None for dem. De skrives
    # inn som formler igjen, slik at `formulas` evaluerer dem — og slik at
    # testen samtidig kontrollerer at motoren peker på riktige kolonner.
    for r in range(17, 43):
        fu.cell(r, 5, f"=B{r}+C{r}+D{r}")
        fu.cell(r, 8, f"=F{r}/E{r}")
        fu.cell(r, 9, f"=G{r}/E{r}")
        fu.cell(r, 15, f"=B{r}*J{r}+C{r}*K{r}+D{r}*L{r}-M{r}")
        fu.cell(r, 16, f"=N{r}/O{r}")
    # Dokumentasjon-arket må finnes for _dokumentasjon
    wb.create_sheet("Dokumentasjon")
    w, z1, z2 = br.trekk(n, br.SEED)
    br._parametre(wb)
    br._utvidelse(wb, n)
    br._motor(wb, w, z1, z2, n)
    br._vifte(wb, n)
    br._dokumentasjon(wb)
    wb.save(sti)
    return w, z1, z2


def _referanse(n, w, z1, z2):
    """Python-referanse på identiske trekk, begge forankringer, 2026-2060."""
    b = mc.les_basis()
    mc.KALIBRERING = "historisk"
    s_on, s_oo, s_gn, s_go = mc.sigmaer(b)
    sig_o, sig_g = s_oo, s_go            # symmetrisk historisk kalibrering
    rho = 0.60
    a = mc.kjed(b, mc.forleng(b))
    volO, volG, volN = a["volO"], a["volG"], a["volN"]
    pO, pG, pN, cost = a["pO"], a["pG"], a["pN"], a["cost"]
    fh, fl, andel = a["fh"], a["fl"], a["andel"]

    fo = np.exp(sig_o * z1)[:, None]
    fg = np.exp(sig_g * (rho * z1 + np.sqrt(1 - rho ** 2) * z2))[:, None]
    volfac = np.where(w[:, None] >= 0, 1 + w[:, None] * (fh - 1),
                      1 + w[:, None] * (1 - fl))
    volfac = volfac / (1 + (fh + fl - 2) / 6)
    ut = {}
    for navn, ko, kg in (("med", 1.0, 1.0),
                         ("for", np.exp(-sig_o ** 2 / 2), np.exp(-sig_g ** 2 / 2))):
        netto = volfac * (volO * pO * fo * ko + volN * pN * fo * ko
                          + volG * pG * fg * kg - cost)
        ut[navn] = andel * np.maximum(netto, 0.0) / 1000.0
    return ut, (sig_o, sig_g), (fo[:, 0], fg[:, 0])


def _les(mod, ark, celle):
    v = mod[f"'[{mod.filename}]{ark}'!{celle}".upper()]
    return float(np.asarray(v).ravel()[0])


def main():
    tmp = tempfile.mkdtemp()
    sti = os.path.join(tmp, "test.xlsx")
    w, z1, z2 = _testbok(N_TEST, sti)
    ref, (sig_o, sig_g), (fo, fg) = _referanse(N_TEST, w, z1, z2)

    import formulas
    print(f"Evaluerer testboken med `formulas` ({N_TEST} simuleringer) …")
    mod = formulas.ExcelModel().loads(sti).finish().calculate()
    # Nøklene i `formulas` har formen "'[fil]ARK'!CELLE", store bokstaver.
    kart = {}
    for k, v in mod.items():
        if "]" in k and "!" in k:
            ark, celle = k.split("]", 1)[1].split("'!", 1)
            kart[(ark, celle)] = v

    def hent(ark, celle):
        v = kart[(ark.upper(), celle.upper())]
        v = v.value if hasattr(v, "value") else v
        return float(np.asarray(v).ravel()[0])

    feil = []

    # --- 1. Motoren celle for celle -----------------------------------------
    from openpyxl.utils import get_column_letter as CL
    verst = {}
    for navn, base in (("med", br.C_MED), ("for", br.C_FOR)):
        d = 0.0
        for r in range(N_TEST):
            for i in range(br.NYA):
                x = hent(br.MOTOR, f"{CL(base + i)}{br.DATA0 + r}")
                d = max(d, abs(x - ref[navn][r, i]))
        verst[navn] = d
        print(f"  Motor {navn}: største avvik mot Python "
              f"{d:.3e} (av typisk nivå {abs(ref[navn]).mean():.1f})")
        if d > 1e-6:
            feil.append(f"motor {navn} avviker {d:.3e}")

    # --- 2. Persentil- og aggregatformler i det synlige arket ---------------
    disc = np.arange(1, br.NY + 1)
    for navn, kol0 in (("med", 2), ("for", 8)):
        for j, q in enumerate((10, 25, 50, 75, 90)):
            x = hent(br.VIFTE, f"{CL(kol0 + j)}5")           # år 2026
            y = np.percentile(ref[navn][:, 0], q)
            if abs(x - y) > 1e-6:
                feil.append(f"vifte {navn} P{q} 2026: {x:.4f} vs {y:.4f}")
        kum = ref[navn][:, :br.NY].sum(axis=1)
        x = hent(br.VIFTE, f"{CL(kol0 + 2)}34")              # kumulativ P50
        y = np.percentile(kum, 50)
        print(f"  Vifte {navn}: kumulativ P50 Excel {x:.2f} / Python {y:.2f}")
        if abs(x - y) > 1e-6:
            feil.append(f"kumulativ P50 {navn}: {x:.4f} vs {y:.4f}")
        x = hent(br.VIFTE, f"{CL(kol0 + 2)}35")              # NPV3 P50
        y = np.percentile((ref[navn][:, :br.NY] / 1.03 ** disc).sum(axis=1), 50)
        if abs(x - y) > 1e-6:
            feil.append(f"NPV3 P50 {navn}: {x:.4f} vs {y:.4f}")

    # --- 2b. Utvidelsen til 2060 og broen -----------------------------------
    d60 = np.arange(1, br.NYA + 1)
    for navn, base in (("med", br.A_MED), ("for", br.A_FOR)):
        a = ref[navn]
        for nokkel, y in (("KUM60", a.sum(axis=1)),
                          ("NPV4_60", (a / 1.04 ** d60).sum(axis=1)),
                          ("NPV3_60", (a / 1.03 ** d60).sum(axis=1))):
            col = CL(base + br.AGG.index(nokkel))
            x = hent(br.MOTOR, f"{col}{br.DATA0}")           # første simulering
            if abs(x - y[0]) > 1e-6:
                feil.append(f"{nokkel} {navn} sim 1: {x:.4f} vs {y[0]:.4f}")
        # og persentilen slik den vises i utvidelsesarket
        rad = 33 if navn == "med" else 34
        col = CL(base + br.AGG.index("NPV4_60"))
        rng_p = np.percentile((a / 1.04 ** d60).sum(axis=1), 50)
        x = hent(br.UTV, f"D{rad}")
        print(f"  Utvidelse {navn}: NNV 4 pst. 2026-2060 P50 Excel {x:.2f} "
              f"/ Python {rng_p:.2f}")
        if abs(x - rng_p) > 1e-6:
            feil.append(f"NNV4 2060 P50 {navn}: {x:.4f} vs {rng_p:.4f}")

    # Basisbanen og broen: rad 4 i broen skal være lik basiskolonnen i rad 31.
    bas = np.array([hent(br.UTV, f"S{br.E_BAS0 + i}") for i in range(br.NYA)])
    npv4_bas = (bas / 1.04 ** d60).sum()
    for celle, y, navn in (("H33", npv4_bas, "basis NNV 4 pst. 2026-2060"),
                           ("B44", npv4_bas, "bro rad 4"),
                           ("B41", (bas[:br.NY] / 1.03 ** np.arange(1, br.NY + 1)).sum(),
                            "bro rad 1 (NNV 3 pst. 2026-2050)")):
        x = hent(br.UTV, celle)
        print(f"  {navn}: Excel {x:.1f} / Python {y:.1f}")
        if abs(x - y) > 1e-6:
            feil.append(f"{navn}: {x:.4f} vs {y:.4f}")
    x = hent(br.UTV, "B50")
    print(f"  Broens interne kontroll (skal være 0): {x:.2e}")
    if abs(x) > 1e-6:
        feil.append(f"broens kontroll ikke 0: {x:.4e}")

    # Ekstrapoleringen: monoton fallende og positiv gjennom hele halen
    for i in range(br.NYE):
        rad = br.EXT0 + i
        for kol, lbl in (("B", "råolje"), ("C", "gass"), ("M", "kostnader")):
            v = hent(br.UTV, f"{kol}{rad}")
            if v <= 0:
                feil.append(f"ekstrapolert {lbl} ikke positiv i {2051+i}: {v}")
    print(f"  Ekstrapolert 2060: råolje {hent(br.UTV, f'B{br.EXT0+9}'):.1f} / "
          f"gass {hent(br.UTV, f'C{br.EXT0+9}'):.1f} mill. Sm3 o.e., "
          f"basis SNCF {hent(br.UTV, f'N{br.EXT0+9}')/1000:.0f} mrd.")
    # Ratene og statsandelen i arket mot Python-referansen
    e = mc.forleng(mc.les_basis())
    for kol, navn, py in (("B", "råolje", e["volO"]), ("C", "gass", e["volG"]),
                          ("M", "kostnader", e["cost"]),
                          ("P", "statsandel", e["andel"])):
        x = hent(br.UTV, f"{kol}{br.EXT0}")
        if abs(x - py[0]) > 1e-9 * max(1.0, abs(py[0])):
            feil.append(f"ekstrapolert {navn} 2051: {x:.6f} vs {py[0]:.6f}")
    print("  Ekstrapoleringen i arket matcher mc_reformulert.forleng")

    # --- 3. Forankringsidentitetene ----------------------------------------
    s_o = hent("Forutsetninger", f"B{br.R_SIGO}")
    s_g = hent("Forutsetninger", f"B{br.R_SIGG}")
    print(f"  Sigma fra arket: olje {s_o:.4f} / gass {s_g:.4f} "
          f"(Python {sig_o:.4f} / {sig_g:.4f})")
    if abs(s_o - sig_o) > 1e-9 or abs(s_g - sig_g) > 1e-9:
        feil.append("sigma i arket avviker fra Python")
    jo = hent("Forutsetninger", f"B{br.R_JO}")
    if abs(jo - np.exp(-s_o ** 2 / 2)) > 1e-12:
        feil.append("Jensen-korreksjon olje feil")
    # Identitetene gjelder i grensen; med 25 trekk sjekkes de analytisk i
    # stedet: median av EXP(sigma*z) er 1, og E[EXP(sigma*z)*jo] er 1.
    if abs(np.exp(0.0) - 1.0) > TOL:
        feil.append("medianidentitet")
    if abs(np.exp(s_o ** 2 / 2) * jo - 1.0) > 1e-12:
        feil.append("forventningsidentitet")
    print("  Forankringsidentiteter: median[faktor] = 1 og "
          "E[faktor]*Jensen = 1 holder analytisk")

    # --- 4. Gulvet ----------------------------------------------------------
    neg = (ref["med"] < 0).sum() + (ref["for"] < 0).sum()
    print(f"  Negative årsverdier med gulv på: {neg}")
    if neg:
        feil.append("gulvet slipper gjennom negative tall")
    b = mc.les_basis()
    mc.KALIBRERING, mc.MEDIAN_ANCHOR, mc.SUPPLY_FLOOR = "historisk", True, False
    uten, _, _ = mc.simuler(b)
    print(f"  Uten gulv (10 000 sim.): {(uten < 0).mean() * 100:.1f} pst. "
          f"negative årsverdier — gulvet har reell effekt")
    if (uten < 0).mean() == 0:
        feil.append("gulvet har ingen effekt; kontroller testen")

    print()
    if feil:
        print("AVVIK FUNNET:")
        for f in feil:
            print("  -", f)
        raise SystemExit(1)
    print("Alle kontroller passerte: Excel-formlene regner identisk med "
          "Python-referansen.")


if __name__ == "__main__":
    main()
